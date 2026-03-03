\
import calendar
import hashlib
import json
import os
from datetime import date, datetime, timedelta, time
from typing import Dict, List, Tuple, Optional

from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from openpyxl import load_workbook

from services.rrhh_db import fetch_all, fetch_one, execute, execute_scalar, call_proc
from services.upload import save_upload
from services.hr_employee_service import employee_can_work_from_home

modulos_bp = Blueprint("modulos", __name__)

# -------------------------
# Helpers
# -------------------------
def _require_admin():
    if not getattr(current_user, "is_admin", False):
        flash("No tienes permisos para acceder a esta sección.", "warning")
        return False
    return True

def _month_range(y: int, m: int) -> Tuple[date, date]:
    first = date(y, m, 1)
    last = date(y, m, calendar.monthrange(y, m)[1])
    return first, last

def _time_to_str(t: Optional[time]) -> str:
    if t is None:
        return ""
    return t.strftime("%H:%M")

def _parse_doc_number(v) -> Optional[str]:
    if v is None:
        return None
    s = (f"{v}").strip()
    if not s:
        return None
    # solo dígitos
    digits = "".join(ch for ch in s if ch.isdigit())
    return digits or None

def _parse_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = (f"{v}").strip()
    if not s:
        return None
    # intenta yyyy-mm-dd o dd/mm/yyyy
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def _parse_time(v) -> Optional[time]:
    if v is None:
        return None
    if isinstance(v, time):
        return v.replace(microsecond=0)
    if isinstance(v, datetime):
        return v.time().replace(microsecond=0)
    s = (f"{v}").strip()
    if not s:
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except Exception:
            pass
    return None

def _diff_minutes(t1: time, t2: time) -> int:
    dt1 = datetime.combine(date(2000, 1, 1), t1)
    dt2 = datetime.combine(date(2000, 1, 1), t2)
    return int((dt2 - dt1).total_seconds() // 60)

def _checksum_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def _att_import_doc_column() -> str:
    # soporta DB vieja (employee_code) o nueva (doc_number)
    col = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS "
        "WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='att_import_row' AND COLUMN_NAME='doc_number'"
    )
    return "doc_number" if col else "employee_code"

# -------------------------
# Dashboard
# -------------------------
@modulos_bp.route("/dashboard")
@login_required
def dashboard():
    puede_trabajo_casa = employee_can_work_from_home(getattr(current_user, "employee_id", None))
    return render_template("dashboard.html", puede_trabajo_casa=puede_trabajo_casa)


@modulos_bp.route("/trabajo-casa/solicitar", methods=["GET", "POST"])
@login_required
def trabajo_casa_solicitar():
    """Solicitud simple de trabajo en casa (por empleado).

    - Visible solo si hr_employee.can_work_from_home = 1
    - Registra en rrhh.wfh_day (si existe)
    """

    employee_id = getattr(current_user, "employee_id", None)
    if employee_id is None:
        flash("Tu usuario no está asociado a un empleado. RRHH debe asignar el perfil.", "warning")
        return redirect(url_for("perfil_pendiente"))

    if not employee_can_work_from_home(employee_id):
        flash("No tienes habilitado el módulo de trabajo en casa.", "warning")
        return redirect(url_for("modulos.dashboard"))

    exists = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='wfh_day'"
    ) is not None
    if not exists:
        flash("Tabla rrhh.wfh_day no existe. Debe crearse por script/migración.", "warning")
        return render_template("modulos/trabajo_casa_solicitar.html", registros=[])

    if request.method == "POST":
        try:
            work_date = datetime.strptime(request.form.get("work_date"), "%Y-%m-%d").date()
            reason = (request.form.get("reason") or "Trabajo en casa").strip()

            created_by = getattr(current_user, "user_db_id", None)
            if created_by is None:
                created_by = getattr(current_user, "user_id", 0)

            execute(
                "MERGE rrhh.wfh_day AS t "
                "USING (SELECT ? AS employee_id, ? AS work_date) AS s "
                "ON (t.employee_id = s.employee_id AND t.work_date = s.work_date) "
                "WHEN MATCHED THEN UPDATE SET is_active=1, reason=? "
                "WHEN NOT MATCHED THEN INSERT (employee_id, work_date, reason, created_by_user_id, is_active) "
                "VALUES (?, ?, ?, ?, 1);",
                (int(employee_id), work_date, reason, int(employee_id), work_date, reason, int(created_by)),
            )
            flash("Solicitud registrada.", "success")
            return redirect(url_for("modulos.trabajo_casa_solicitar"))
        except Exception as ex:
            flash(f"No se pudo registrar la solicitud: {ex}", "danger")

    registros = fetch_all(
        "SELECT work_date, reason "
        "FROM rrhh.wfh_day "
        "WHERE employee_id=? AND is_active=1 "
        "ORDER BY work_date DESC",
        (int(employee_id),),
    )
    return render_template("modulos/trabajo_casa_solicitar.html", registros=registros)

# -------------------------
# Turnos (asignación de horario)
# -------------------------
@modulos_bp.route("/turnos")
@login_required
def turnos():
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    today = date.today()
    rows = fetch_all(
        "SELECT e.employee_id, e.doc_number, e.first_name, e.last_name, "
        "sd.shift_code, sd.start_time, sd.end_time "
        "FROM rrhh.hr_employee e "
        "OUTER APPLY ("
        "  SELECT TOP (1) sa.shift_id "
        "  FROM rrhh.shift_assignment sa "
        "  WHERE sa.employee_id = e.employee_id "
        "    AND sa.valid_from <= ? "
        "    AND (sa.valid_to IS NULL OR sa.valid_to >= ?) "
        "  ORDER BY sa.valid_from DESC"
        ") x "
        "LEFT JOIN rrhh.shift_definition sd ON sd.shift_id = x.shift_id "
        "WHERE e.is_active = 1 "
        "ORDER BY e.last_name, e.first_name",
        (today, today),
    )

    shifts = fetch_all(
        "SELECT shift_id, shift_code, start_time, end_time FROM rrhh.shift_definition WHERE is_active=1 ORDER BY shift_id"
    )
    return render_template("modulos/turnos.html", empleados=rows, shifts=shifts)

@modulos_bp.route("/turnos/asignar", methods=["GET", "POST"])
@login_required
def turnos_asignar():
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    empleados = fetch_all(
        "SELECT employee_id, doc_number, first_name, last_name FROM rrhh.hr_employee WHERE is_active=1 ORDER BY last_name, first_name"
    )
    shifts = fetch_all(
        "SELECT shift_id, shift_code, start_time, end_time FROM rrhh.shift_definition WHERE is_active=1 ORDER BY shift_id"
    )

    if request.method == "POST":
        employee_id = int(request.form.get("employee_id"))
        shift_id = int(request.form.get("shift_id"))
        valid_from = request.form.get("valid_from")
        valid_to = request.form.get("valid_to") or None

        try:
            vf = datetime.strptime(valid_from, "%Y-%m-%d").date()
            vt = datetime.strptime(valid_to, "%Y-%m-%d").date() if valid_to else None
            call_proc(
                "rrhh.sp_set_shift_assignment",
                [employee_id, shift_id, vf, vt, current_user.user_id, "Asignación manual"],
            )
            flash("Turno asignado correctamente.", "success")
            return redirect(url_for("modulos.turnos"))
        except Exception as ex:
            flash(f"No se pudo asignar el turno: {ex}", "danger")

    return render_template("modulos/turnos_asignar.html", empleados=empleados, shifts=shifts)

# -------------------------
# Asistencia (import + vista)
# -------------------------
def _parse_attendance_excel(file_path: str) -> List[Dict]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # detectar cabecera (primera fila)
    header = []
    for cell in ws[1]:
        header.append(str(cell.value).strip().lower() if cell.value is not None else "")
    header_map = {name: idx for idx, name in enumerate(header)}

    def find_col(patterns):
        for name, idx in header_map.items():
            for p in patterns:
                if p in name:
                    return idx
        return None

    col_doc = find_col(["cedula", "cédula", "documento", "doc", "identific", "cc", "c.c", "dni", "id"])
    col_date = find_col(["fecha", "date"])
    col_in = find_col(["entrada", "first", "ingreso", "in"])
    col_out = find_col(["salida", "last", "egreso", "out"])
    col_min = find_col(["min", "minutos", "total", "tiempo", "duracion", "duración"])

    if col_doc is None or col_date is None:
        raise ValueError("No se encontró columna de Cédula/Documento y/o Fecha en el Excel (revisa encabezados).")

    grouped: Dict[Tuple[str, date], Dict] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        doc = _parse_doc_number(r[col_doc] if col_doc < len(r) else None)
        wdate = _parse_date(r[col_date] if col_date < len(r) else None)
        if not doc or not wdate:
            continue

        fin = _parse_time(r[col_in] if (col_in is not None and col_in < len(r)) else None)
        fout = _parse_time(r[col_out] if (col_out is not None and col_out < len(r)) else None)
        mins = None
        if col_min is not None and col_min < len(r):
            try:
                mins = int(float(r[col_min])) if r[col_min] is not None else None
            except Exception:
                mins = None

        key = (doc, wdate)
        entry = grouped.get(key)
        raw = {"row": [("" if x is None else f"{x}") for x in r]}

        if entry is None:
            grouped[key] = {
                "doc_number": doc,
                "work_date": wdate,
                "first_in": fin,
                "last_out": fout,
                "total_minutes": mins,
                "raw": [raw],
            }
        else:
            if fin and (entry["first_in"] is None or fin < entry["first_in"]):
                entry["first_in"] = fin
            if fout and (entry["last_out"] is None or fout > entry["last_out"]):
                entry["last_out"] = fout
            # si viene minutes, dejamos el máximo
            if mins is not None:
                if entry["total_minutes"] is None or mins > entry["total_minutes"]:
                    entry["total_minutes"] = mins
            entry["raw"].append(raw)

    rows = []
    for v in grouped.values():
        if v["total_minutes"] is None and v["first_in"] and v["last_out"]:
            try:
                v["total_minutes"] = max(0, _diff_minutes(v["first_in"], v["last_out"]))
            except Exception:
                v["total_minutes"] = None
        v["raw_text"] = json.dumps(v["raw"], ensure_ascii=False)
        rows.append(v)

    return rows

@modulos_bp.route("/asistencia")
@login_required
def asistencia():
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    y = int(request.args.get("year", date.today().year))
    m = int(request.args.get("month", date.today().month))
    d_from, d_to = _month_range(y, m)

    # fechas del mes
    days = [d_from + timedelta(days=i) for i in range((d_to - d_from).days + 1)]

    empleados = fetch_all(
        "SELECT employee_id, doc_number, first_name, last_name FROM rrhh.hr_employee WHERE is_active=1 ORDER BY last_name, first_name"
    )

    att_rows = fetch_all(
        "SELECT employee_id, work_date, first_in, last_out, total_minutes, has_manual_override "
        "FROM rrhh.vw_attendance_effective "
        "WHERE work_date BETWEEN ? AND ?",
        (d_from, d_to),
    )
    att_map = {(r.employee_id, r.work_date): r for r in att_rows}

    # WFH (si existe)
    wfh_exists = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='wfh_day'"
    ) is not None
    wfh_map = set()
    if wfh_exists:
        wfh_rows = fetch_all(
            "SELECT employee_id, work_date FROM rrhh.wfh_day WHERE is_active=1 AND work_date BETWEEN ? AND ?",
            (d_from, d_to),
        )
        wfh_map = {(r.employee_id, r.work_date) for r in wfh_rows}

    # time reduction (si existe y está aprobado)
    tr_exists = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='time_reduction_request'"
    ) is not None
    reduction_map = {}
    if tr_exists:
        tr_rows = fetch_all(
            "SELECT r.employee_id, t.reduction_date, t.minutes "
            "FROM rrhh.wf_request r "
            "JOIN rrhh.time_reduction_request t ON t.request_id = r.request_id "
            "WHERE r.request_type = 'TIME_REDUCTION' AND r.status = 'APPROVED' "
            "AND t.reduction_date BETWEEN ? AND ?",
            (d_from, d_to),
        )
        for rr in tr_rows:
            reduction_map[(rr.employee_id, rr.reduction_date)] = int(rr.minutes or 60)

    # shift assignments (traer los que cruzan el mes)
    shifts = fetch_all(
        "SELECT sa.employee_id, sa.valid_from, sa.valid_to, sd.shift_code, sd.start_time, sd.end_time "
        "FROM rrhh.shift_assignment sa "
        "JOIN rrhh.shift_definition sd ON sd.shift_id = sa.shift_id "
        "WHERE sa.valid_from <= ? AND (sa.valid_to IS NULL OR sa.valid_to >= ?) "
        "ORDER BY sa.employee_id, sa.valid_from",
        (d_to, d_from),
    )
    shifts_by_emp: Dict[int, List] = {}
    for s in shifts:
        shifts_by_emp.setdefault(s.employee_id, []).append(s)

    def shift_for(emp_id: int, day: date):
        lst = shifts_by_emp.get(emp_id, [])
        best = None
        for s in lst:
            if s.valid_from <= day and (s.valid_to is None or s.valid_to >= day):
                best = s
        return best

    matrix = []
    for e in empleados:
        row_days = []
        summary = {"ok": 0, "faltante": 0, "incompleto": 0, "casa": 0, "sin_turno": 0}
        for d in days:
            if d.weekday() >= 5:  # sábado/domingo
                row_days.append({"status": "—", "title": "Fin de semana", "link": None})
                continue

            sh = shift_for(e.employee_id, d)
            if not sh:
                summary["sin_turno"] += 1
                row_days.append({"status": "ST", "title": "Sin turno asignado", "link": None})
                continue

            required = _diff_minutes(sh.start_time, sh.end_time)
            required -= reduction_map.get((e.employee_id, d), 0)
            required = max(0, required)

            att = att_map.get((e.employee_id, d))
            if att is None or att.total_minutes is None:
                if (e.employee_id, d) in wfh_map:
                    summary["casa"] += 1
                    row_days.append({"status": "CASA", "title": f"Trabajo en casa. Requerido {required} min", "link": url_for("modulos.asistencia_ajuste", employee_id=e.employee_id, work_date=d.isoformat())})
                else:
                    summary["faltante"] += 1
                    row_days.append({"status": "F", "title": f"Faltante. Requerido {required} min", "link": url_for("modulos.asistencia_ajuste", employee_id=e.employee_id, work_date=d.isoformat())})
                continue

            mins = int(att.total_minutes or 0)
            ok = mins >= required
            if ok:
                summary["ok"] += 1
                st = "OK*" if att.has_manual_override else "OK"
                row_days.append({"status": st, "title": f"{mins} min (req {required}). In {_time_to_str(att.first_in)} / Out {_time_to_str(att.last_out)}", "link": url_for("modulos.asistencia_ajuste", employee_id=e.employee_id, work_date=d.isoformat())})
            else:
                summary["incompleto"] += 1
                row_days.append({"status": "I", "title": f"Incompleto: {mins} min (req {required}). In {_time_to_str(att.first_in)} / Out {_time_to_str(att.last_out)}", "link": url_for("modulos.asistencia_ajuste", employee_id=e.employee_id, work_date=d.isoformat())})

        # turno actual para mostrar
        sh_today = shift_for(e.employee_id, date.today())
        shift_label = sh_today.shift_code if sh_today else "—"
        matrix.append({
            "employee_id": e.employee_id,
            "doc_number": e.doc_number,
            "name": f"{e.first_name} {e.last_name}",
            "shift": shift_label,
            "days": row_days,
            "summary": summary,
        })

    return render_template("modulos/asistencia.html", year=y, month=m, days=days, matrix=matrix)

@modulos_bp.route("/asistencia/cargar", methods=["GET", "POST"])
@login_required
def asistencia_cargar():
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    if request.method == "POST":
        year_no = int(request.form.get("year_no"))
        month_no = int(request.form.get("month_no"))
        f = request.files.get("file")

        if not f or not f.filename:
            flash("Debes seleccionar un archivo.", "warning")
            return render_template("modulos/asistencia_cargar.html", year_no=year_no, month_no=month_no)

        try:
            meta = save_upload(f, prefix="asistencia")
            checksum = _checksum_file(meta["storage_path"])

            # registrar archivo
            execute(
                "INSERT INTO rrhh.sys_attachment(file_name, mime_type, size_bytes, storage_path, uploaded_by_user_id) "
                "VALUES (?, ?, ?, ?, ?)",
                (meta["file_name"], f.mimetype or "application/octet-stream", meta["size_bytes"], meta["storage_path"], current_user.user_id),
            )
            file_id = execute_scalar("SELECT SCOPE_IDENTITY()")

            # batch
            execute(
                "INSERT INTO rrhh.att_import_batch(year_no, month_no, file_name, file_id, checksum, uploaded_by_user_id) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (year_no, month_no, meta["file_name"], file_id, checksum, current_user.user_id),
            )
            batch_id = int(execute_scalar("SELECT SCOPE_IDENTITY()"))

            # parse excel (todos los registros)
            parsed = _parse_attendance_excel(meta["storage_path"])
            doc_col = _att_import_doc_column()

            # insert rows (OK / ERROR si no existe empleado)
            for r in parsed:
                # validar empleado existe
                emp = fetch_one("SELECT employee_id FROM rrhh.hr_employee WHERE doc_number = ? AND is_active=1", (r["doc_number"],))
                if not emp:
                    execute(
                        f"INSERT INTO rrhh.att_import_row(batch_id, {doc_col}, work_date, first_in, last_out, total_minutes, raw_text, load_status, error_message) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, 'ERROR', ?)",
                        (batch_id, r["doc_number"], r["work_date"], r["first_in"], r["last_out"], r["total_minutes"], r["raw_text"], "Empleado no existe (doc_number no encontrado)"),
                    )
                else:
                    execute(
                        f"INSERT INTO rrhh.att_import_row(batch_id, {doc_col}, work_date, first_in, last_out, total_minutes, raw_text, load_status, error_message) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, 'OK', NULL)",
                        (batch_id, r["doc_number"], r["work_date"], r["first_in"], r["last_out"], r["total_minutes"], r["raw_text"]),
                    )

            # aplicar a att_day
            call_proc("rrhh.sp_apply_attendance_batch", [batch_id, current_user.user_id])

            flash(f"Asistencia cargada y aplicada. Batch #{batch_id}.", "success")
            return redirect(url_for("modulos.asistencia", year=year_no, month=month_no))

        except Exception as ex:
            flash(f"Error cargando asistencia: {ex}", "danger")

    today = date.today()
    return render_template("modulos/asistencia_cargar.html", year_no=today.year, month_no=today.month)

@modulos_bp.route("/asistencia/ajuste", methods=["GET", "POST"])
@login_required
def asistencia_ajuste():
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    employee_id = int(request.values.get("employee_id"))
    work_date_s = request.values.get("work_date")
    work_date = datetime.strptime(work_date_s, "%Y-%m-%d").date()

    emp = fetch_one("SELECT employee_id, doc_number, first_name, last_name FROM rrhh.hr_employee WHERE employee_id=?", (employee_id,))
    if not emp:
        flash("Empleado no encontrado.", "warning")
        return redirect(url_for("modulos.asistencia"))

    # asistencia efectiva
    att = fetch_one(
        "SELECT employee_id, work_date, first_in, last_out, total_minutes, has_manual_override "
        "FROM rrhh.vw_attendance_effective WHERE employee_id=? AND work_date=?",
        (employee_id, work_date),
    )

    # turno del día y requerido
    sh = fetch_one(
        "SELECT TOP (1) sd.shift_code, sd.start_time, sd.end_time "
        "FROM rrhh.shift_assignment sa "
        "JOIN rrhh.shift_definition sd ON sd.shift_id = sa.shift_id "
        "WHERE sa.employee_id=? AND sa.valid_from <= ? AND (sa.valid_to IS NULL OR sa.valid_to >= ?) "
        "ORDER BY sa.valid_from DESC",
        (employee_id, work_date, work_date),
    )
    required = _diff_minutes(sh.start_time, sh.end_time) if sh else None

    if request.method == "POST":
        reason = (request.form.get("reason") or "").strip()
        comment = (request.form.get("comment") or "").strip() or None
        first_in = _parse_time(request.form.get("first_in") or None)
        last_out = _parse_time(request.form.get("last_out") or None)
        total_minutes = request.form.get("total_minutes")
        total_minutes = int(total_minutes) if total_minutes not in (None, "",) else None

        # quick action: cumplir
        if request.form.get("action") == "cumplio":
            if required is None:
                flash("No hay turno asignado para calcular requerido.", "warning")
                return redirect(url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=work_date.isoformat()))
            total_minutes = required
            if not reason:
                reason = "NOVEDAD: cumplimiento (salida temprana u otra)"

        if not reason:
            flash("Debes indicar un motivo (reason).", "warning")
            return redirect(url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=work_date.isoformat()))

        try:
            # desactivar overrides anteriores
            execute(
                "UPDATE rrhh.att_manual_override SET is_active=0 "
                "WHERE employee_id=? AND work_date=? AND is_active=1",
                (employee_id, work_date),
            )
            execute(
                "INSERT INTO rrhh.att_manual_override(employee_id, work_date, first_in, last_out, total_minutes, reason, comment, created_by_user_id, is_active) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)",
                (employee_id, work_date, first_in, last_out, total_minutes, reason, comment, current_user.user_id),
            )
            flash("Ajuste/Novedad guardado.", "success")
            return redirect(url_for("modulos.asistencia", year=work_date.year, month=work_date.month))
        except Exception as ex:
            flash(f"No se pudo guardar el ajuste: {ex}", "danger")

    return render_template(
        "modulos/asistencia_ajuste.html",
        emp=emp,
        work_date=work_date,
        att=att,
        shift=sh,
        required_minutes=required,
    )

# -------------------------
# Trabajo en casa (mínimo)
# -------------------------
@modulos_bp.route("/trabajo-casa")
@login_required
def trabajo_casa():
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    exists = fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='wfh_day'") is not None
    if not exists:
        flash("Tabla rrhh.wfh_day no existe. Aplica el script SQL de desarrollo para crearla.", "warning")
        return render_template("modulos/trabajo_casa.html", registros=[], empleados=[])

    registros = fetch_all(
        "SELECT w.employee_id, w.work_date, w.reason, e.doc_number, e.first_name, e.last_name "
        "FROM rrhh.wfh_day w JOIN rrhh.hr_employee e ON e.employee_id = w.employee_id "
        "WHERE w.is_active=1 ORDER BY w.work_date DESC"
    )
    empleados = fetch_all(
        "SELECT employee_id, doc_number, first_name, last_name FROM rrhh.hr_employee WHERE is_active=1 ORDER BY last_name, first_name"
    )
    return render_template("modulos/trabajo_casa.html", registros=registros, empleados=empleados)

@modulos_bp.route("/trabajo-casa/nuevo", methods=["POST"])
@login_required
def trabajo_casa_nuevo():
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    employee_id = int(request.form.get("employee_id"))
    work_date = datetime.strptime(request.form.get("work_date"), "%Y-%m-%d").date()
    reason = (request.form.get("reason") or "Trabajo en casa").strip()

    try:
        execute(
            "MERGE rrhh.wfh_day AS t "
            "USING (SELECT ? AS employee_id, ? AS work_date) AS s "
            "ON (t.employee_id = s.employee_id AND t.work_date = s.work_date) "
            "WHEN MATCHED THEN UPDATE SET is_active=1, reason=? "
            "WHEN NOT MATCHED THEN INSERT (employee_id, work_date, reason, created_by_user_id, is_active) "
            "VALUES (?, ?, ?, ?, 1);",
            (employee_id, work_date, reason, employee_id, work_date, reason, current_user.user_id),
        )
        flash("Día de trabajo en casa registrado.", "success")
    except Exception as ex:
        flash(f"No se pudo registrar: {ex}", "danger")

    return redirect(url_for("modulos.trabajo_casa"))
