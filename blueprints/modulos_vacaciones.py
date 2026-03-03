from __future__ import annotations

from datetime import date, datetime, timedelta

from flask import render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user

from services.rrhh_db import fetch_all, fetch_one, execute, execute_scalar, call_proc
from services.upload import save_upload
from services.rrhh_security import ROLE_ADMIN, ROLE_RRHH

from .modulos import modulos_bp
from .modulos_common import _is_admin_or_rrhh, _checksum_file, _parse_doc_number


REQUEST_TYPE_VACACIONES = "VACACIONES"


# -----------------------------------------------------------------------------
# Helpers (DB / roles)
# -----------------------------------------------------------------------------


def _table_exists(schema: str, table: str) -> bool:
    row = fetch_one(
        "SELECT 1 AS ok FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA=? AND TABLE_NAME=?",
        (schema, table),
    )
    return row is not None


def _has_col(schema: str, table: str, col: str) -> bool:
    row = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS "
        "WHERE TABLE_SCHEMA=? AND TABLE_NAME=? AND COLUMN_NAME=?",
        (schema, table, col),
    )
    return row is not None


def _roles() -> set[str]:
    r = getattr(current_user, "roles", None) or []
    try:
        return set(r)
    except Exception:
        return set(list(r))


def _is_admin() -> bool:
    roles = _roles()
    return bool(
        getattr(current_user, "is_admin", False)
        or (ROLE_ADMIN in roles)
        or ("ADMINISTRADOR" in roles)
        or ("Administrador" in roles)
    )


def _is_rrhh() -> bool:
    roles = _roles()
    return bool((ROLE_RRHH in roles) or ("RRHH" in roles) or ("Recursos Humanos" in roles))


def _vacaciones_ready() -> bool:
    return _table_exists("rrhh", "vacation_balance") and _table_exists(
        "rrhh", "vacation_request_detail"
    )


# -----------------------------------------------------------------------------
# Helpers (business rules)
# -----------------------------------------------------------------------------


def _get_holidays_between(d_from: date, d_to: date) -> set[date]:
    """Retorna festivos activos entre d_from y d_to (inclusive)."""
    if not _table_exists("rrhh", "hr_holiday"):
        return set()

    sql = "SELECT holiday_date FROM rrhh.hr_holiday WHERE holiday_date BETWEEN ? AND ?"
    if _has_col("rrhh", "hr_holiday", "is_active"):
        sql += " AND is_active=1"

    rows = fetch_all(sql, (d_from, d_to))
    return {r.holiday_date for r in (rows or []) if getattr(r, "holiday_date", None)}


def _count_business_days(d_from: date, d_to: date) -> int:
    """Cuenta días hábiles (Lun-Vie) excluyendo festivos activos."""
    if d_to < d_from:
        return 0

    holidays = _get_holidays_between(d_from, d_to)
    cur = d_from
    c = 0
    while cur <= d_to:
        if cur.weekday() < 5 and cur not in holidays:
            c += 1
        cur += timedelta(days=1)
    return c


def _get_balance(employee_id: int) -> dict | None:
    if not employee_id or not _table_exists("rrhh", "vacation_balance"):
        return None

    row = fetch_one(
        "SELECT employee_id, available_days, used_days, as_of_date, updated_at "
        "FROM rrhh.vacation_balance WHERE employee_id=?",
        (int(employee_id),),
    )
    if not row:
        return None

    available = float(getattr(row, "available_days", 0) or 0)
    used = float(getattr(row, "used_days", 0) or 0)
    total = max(available + used, 0)

    return {
        "employee_id": int(employee_id),
        "available": available,
        "used": used,
        "total": total,
        "as_of_date": getattr(row, "as_of_date", None),
        "updated_at": getattr(row, "updated_at", None),
    }


def _has_overlapping_request(employee_id: int, start_date: date, end_date: date) -> bool:
    """Evita solapar vacaciones SUBMITTED/APPROVED del mismo empleado."""
    if not employee_id:
        return False

    row = fetch_one(
        "SELECT TOP 1 1 AS ok "
        "FROM rrhh.wf_request r "
        "JOIN rrhh.vacation_request_detail d ON d.request_id=r.request_id "
        "WHERE r.request_type=? "
        "  AND r.employee_id=? "
        "  AND r.status IN ('SUBMITTED','APPROVED') "
        "  AND d.start_date <= ? "
        "  AND d.end_date >= ?",
        (REQUEST_TYPE_VACACIONES, int(employee_id), end_date, start_date),
    )
    return row is not None


def _approve_final_step_and_apply_balance(
    *,
    step_id: int,
    request_id: int,
    step_no: int,
    actor_user_id: int,
    comment: str | None,
) -> None:
    """Aprueba el último paso, cierra el request y descuenta saldo (TODO en una transacción)."""

    sql = (
        "BEGIN TRAN\n"
        "  -- 1) Aprobar paso (debe seguir PENDING)\n"
        "  UPDATE rrhh.wf_request_step\n"
        "     SET status='APPROVED', acted_at=GETDATE(), comment=?\n"
        "   WHERE step_id=? AND status='PENDING';\n"
        "  IF @@ROWCOUNT = 0\n"
        "  BEGIN\n"
        "    ROLLBACK TRAN;\n"
        "    RAISERROR('El paso no está pendiente o ya fue gestionado.',16,1);\n"
        "    RETURN;\n"
        "  END\n"
        "\n"
        "  -- 2) Trazabilidad\n"
        "  INSERT INTO rrhh.wf_action(request_id, step_no, actor_user_id, action, comment)\n"
        "  VALUES (?, ?, ?, 'APPROVE', ?);\n"
        "\n"
        "  -- 3) Cerrar request\n"
        "  UPDATE rrhh.wf_request\n"
        "     SET status='APPROVED', closed_at=GETDATE()\n"
        "   WHERE request_id=?;\n"
        "\n"
        "  -- 4) Descontar saldo\n"
        "  DECLARE @emp_id INT, @total DECIMAL(10,2);\n"
        "  SELECT @emp_id = employee_id FROM rrhh.wf_request WHERE request_id=?;\n"
        "  SELECT @total = days_total FROM rrhh.vacation_request_detail WHERE request_id=?;\n"
        "  IF @emp_id IS NULL OR @total IS NULL\n"
        "  BEGIN\n"
        "    ROLLBACK TRAN;\n"
        "    RAISERROR('No se pudo aplicar saldo: solicitud incompleta.',16,1);\n"
        "    RETURN;\n"
        "  END\n"
        "\n"
        "  IF EXISTS (SELECT 1 FROM rrhh.vacation_request_detail WHERE request_id=? AND balance_applied_at IS NOT NULL)\n"
        "  BEGIN\n"
        "    ROLLBACK TRAN;\n"
        "    RAISERROR('El saldo ya fue aplicado para esta solicitud.',16,1);\n"
        "    RETURN;\n"
        "  END\n"
        "\n"
        "  UPDATE rrhh.vacation_balance\n"
        "     SET available_days = available_days - @total,\n"
        "         used_days = used_days + @total,\n"
        "         updated_by_user_id = ?,\n"
        "         updated_at = GETDATE()\n"
        "   WHERE employee_id = @emp_id AND available_days >= @total;\n"
        "\n"
        "  IF @@ROWCOUNT = 0\n"
        "  BEGIN\n"
        "    ROLLBACK TRAN;\n"
        "    RAISERROR('Saldo insuficiente o no existe balance para el empleado.',16,1);\n"
        "    RETURN;\n"
        "  END\n"
        "\n"
        "  UPDATE rrhh.vacation_request_detail\n"
        "     SET balance_applied_at = GETDATE(),\n"
        "         balance_applied_by_user_id = ?\n"
        "   WHERE request_id=? AND balance_applied_at IS NULL;\n"
        "\n"
        "  IF @@ROWCOUNT = 0\n"
        "  BEGIN\n"
        "    ROLLBACK TRAN;\n"
        "    RAISERROR('No se pudo marcar la solicitud como aplicada (posible doble aplicación).',16,1);\n"
        "    RETURN;\n"
        "  END\n"
        "COMMIT TRAN"
    )

    execute(
        sql,
        (
            comment,
            int(step_id),
            int(request_id),
            int(step_no),
            int(actor_user_id),
            comment,
            int(request_id),
            int(request_id),
            int(request_id),
            int(request_id),
            int(actor_user_id),
            int(actor_user_id),
            int(request_id),
        ),
    )


def _reject_step_and_close_request(
    *,
    step_id: int,
    request_id: int,
    step_no: int,
    actor_user_id: int,
    comment: str | None,
) -> None:
    """Rechaza un paso, cierra el request y omite los pasos restantes (transaccional)."""

    sql = (
        "BEGIN TRAN\n"
        "  UPDATE rrhh.wf_request_step\n"
        "     SET status='REJECTED', acted_at=GETDATE(), comment=?\n"
        "   WHERE step_id=? AND status='PENDING';\n"
        "  IF @@ROWCOUNT = 0\n"
        "  BEGIN\n"
        "    ROLLBACK TRAN;\n"
        "    RAISERROR('El paso no está pendiente o ya fue gestionado.',16,1);\n"
        "    RETURN;\n"
        "  END\n"
        "\n"
        "  INSERT INTO rrhh.wf_action(request_id, step_no, actor_user_id, action, comment)\n"
        "  VALUES (?, ?, ?, 'REJECT', ?);\n"
        "\n"
        "  UPDATE rrhh.wf_request\n"
        "     SET status='REJECTED', closed_at=GETDATE()\n"
        "   WHERE request_id=?;\n"
        "\n"
        "  UPDATE rrhh.wf_request_step\n"
        "     SET status='SKIPPED', acted_at=GETDATE()\n"
        "   WHERE request_id=? AND status='PENDING';\n"
        "COMMIT TRAN"
    )

    execute(
        sql,
        (
            comment,
            int(step_id),
            int(request_id),
            int(step_no),
            int(actor_user_id),
            comment,
            int(request_id),
            int(request_id),
        ),
    )


# -----------------------------------------------------------------------------
# Excel parser (Libro de Vacaciones)
# -----------------------------------------------------------------------------


def _parse_vacaciones_excel(file_path: str) -> list[dict]:
    """Parsea el 'Libro de Vacaciones' (formato nómina).

    Retorna lista de dicts:
      {doc_number, employee_name, used_days, available_days}

    Estrategia:
    - El archivo no viene en formato tabla plana.
    - Se recorre por filas.
    - Cuando se detecta una fila 'Total Empleado' se toma el acumulado y se asocia
      a la última cédula encontrada.
    """

    try:
        from openpyxl import load_workbook  # lazy import
    except Exception as ex:
        raise RuntimeError(
            "Dependencia faltante: openpyxl. Instálala para poder cargar el Libro de Vacaciones."
        ) from ex

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    records: list[dict] = []

    current_doc: str | None = None
    current_name: str | None = None

    def _is_total_row(vals) -> bool:
        for v in vals:
            if isinstance(v, str):
                s = v.strip().lower()
                if s.startswith("total") and "empleado" in s:
                    return True
        return False

    def _to_float(v) -> float | None:
        if v is None or v == "":
            return None
        try:
            return float(v)
        except Exception:
            try:
                return float(f"{v}".strip().replace(",", "."))
            except Exception:
                return None

    for row in ws.iter_rows(values_only=True):
        vals = list(row or [])

        # cédula suele venir en la col 3 (index 2). Igual la normalizamos a dígitos.
        doc = None
        if len(vals) >= 3:
            doc = _parse_doc_number(vals[2])

        if doc:
            current_doc = doc
            name = None
            if len(vals) >= 4 and vals[3] is not None:
                name = str(vals[3]).strip()
            current_name = name or current_name

        if _is_total_row(vals) and current_doc:
            # En este formato, las últimas 2 columnas numéricas suelen ser:
            # - Días (disfrutados / gastados)
            # - Pendientes (disponible)
            used = None
            available = None

            # Preferimos col 19 y 20 si existen (index 18,19)
            if len(vals) >= 20:
                used = _to_float(vals[18])
                available = _to_float(vals[19])

            # fallback: tomar últimos 2 números
            if used is None or available is None:
                nums = [_to_float(v) for v in vals if _to_float(v) is not None]
                nums = [n for n in nums if n is not None]
                if len(nums) >= 2:
                    used = nums[-2]
                    available = nums[-1]

            if used is None:
                used = 0.0
            if available is None:
                available = 0.0

            records.append(
                {
                    "doc_number": current_doc,
                    "employee_name": (current_name or "").strip() or None,
                    "used_days": float(used),
                    "available_days": float(available),
                }
            )

            current_doc = None
            current_name = None

    # Deduplicar por doc (si aparece repetido) quedándonos con la última fila total
    by_doc: dict[str, dict] = {}
    for r in records:
        d = r.get("doc_number")
        if d:
            by_doc[d] = r

    return list(by_doc.values())


# -----------------------------------------------------------------------------
# Routes
# -----------------------------------------------------------------------------


@modulos_bp.route("/vacaciones")
@login_required
def vacaciones():
    """Vista del empleado: saldo + solicitudes."""

    if not _vacaciones_ready():
        flash(
            "Configuración pendiente: no se encontró rrhh.vacation_balance / rrhh.vacation_request_detail en la BD.",
            "warning",
        )
        return render_template(
            "modulos/vacaciones.html",
            ready=False,
            balance=None,
            requests=[],
            today=date.today(),
        )

    emp_id = getattr(current_user, "employee_id", None)
    if not emp_id:
        flash("Tu usuario no está asociado a un empleado. Contacta a RRHH.", "warning")
        return render_template(
            "modulos/vacaciones.html",
            ready=True,
            balance=None,
            requests=[],
            today=date.today(),
        )

    balance = _get_balance(int(emp_id))

    reqs = fetch_all(
        "SELECT TOP 100 "
        "  r.request_id, r.status, r.created_at, r.submitted_at, r.closed_at, "
        "  d.start_date, d.end_date, d.days_enjoyed, d.days_paid, d.days_total, d.notes, "
        "  d.balance_applied_at "
        "FROM rrhh.wf_request r "
        "JOIN rrhh.vacation_request_detail d ON d.request_id=r.request_id "
        "WHERE r.request_type=? AND r.employee_id=? "
        "ORDER BY r.created_at DESC",
        (REQUEST_TYPE_VACACIONES, int(emp_id)),
    )

    return render_template(
        "modulos/vacaciones.html",
        ready=True,
        balance=balance,
        requests=reqs or [],
        today=date.today(),
    )


@modulos_bp.route("/vacaciones/solicitar", methods=["POST"])
@login_required
def vacaciones_solicitar():
    if not _vacaciones_ready():
        flash("Módulo no configurado en BD. Ejecuta el patch de vacaciones.", "warning")
        return redirect(url_for("modulos.vacaciones"))

    emp_id = getattr(current_user, "employee_id", None)
    if not emp_id:
        flash("Tu usuario no está asociado a un empleado. Contacta a RRHH.", "warning")
        return redirect(url_for("modulos.vacaciones"))

    # Parse form
    start_s = (request.form.get("start_date") or "").strip()
    end_s = (request.form.get("end_date") or "").strip()
    notes = (request.form.get("notes") or "").strip() or None

    paid_raw = (request.form.get("days_paid") or "").strip()
    try:
        days_paid = int(paid_raw) if paid_raw else 0
    except Exception:
        flash("Días pagas inválidos. Usa un número entero.", "warning")
        return redirect(url_for("modulos.vacaciones"))

    def _parse_iso(d: str) -> date | None:
        try:
            return datetime.strptime(d, "%Y-%m-%d").date()
        except Exception:
            return None

    d_from = _parse_iso(start_s)
    d_to = _parse_iso(end_s)

    if not d_from or not d_to:
        flash("Debes seleccionar fecha inicio y fin.", "warning")
        return redirect(url_for("modulos.vacaciones"))

    if d_to < d_from:
        flash("La fecha fin no puede ser menor a la fecha inicio.", "warning")
        return redirect(url_for("modulos.vacaciones"))

    if days_paid < 0:
        flash("Días pagas no puede ser negativo.", "warning")
        return redirect(url_for("modulos.vacaciones"))

    days_enjoyed = _count_business_days(d_from, d_to)
    if days_enjoyed <= 0:
        flash("El periodo seleccionado no tiene días hábiles para disfrutar.", "warning")
        return redirect(url_for("modulos.vacaciones"))

    if days_paid > days_enjoyed:
        flash(
            "Los días pagas no pueden ser superiores a los días de disfrute (hábiles) del periodo.",
            "warning",
        )
        return redirect(url_for("modulos.vacaciones"))

    total_days = float(days_enjoyed + days_paid)

    bal = _get_balance(int(emp_id))
    if not bal:
        flash("No tienes saldo cargado. RRHH debe cargar el Libro de Vacaciones.", "warning")
        return redirect(url_for("modulos.vacaciones"))

    if float(bal.get("available", 0) or 0) < total_days:
        flash(
            f"Saldo insuficiente. Disponible: {bal.get('available', 0)} días. Estás solicitando: {total_days}.",
            "warning",
        )
        return redirect(url_for("modulos.vacaciones"))

    if _has_overlapping_request(int(emp_id), d_from, d_to):
        flash("Ya tienes una solicitud de vacaciones que se cruza con ese periodo.", "warning")
        return redirect(url_for("modulos.vacaciones"))

    # Crear request workflow
    request_id = execute_scalar(
        "INSERT INTO rrhh.wf_request(request_type, employee_id, created_by_user_id) "
        "OUTPUT INSERTED.request_id VALUES (?, ?, ?)",
        (REQUEST_TYPE_VACACIONES, int(emp_id), int(current_user.user_id)),
    )

    if not request_id:
        flash("No se pudo crear la solicitud.", "danger")
        return redirect(url_for("modulos.vacaciones"))

    execute(
        "INSERT INTO rrhh.vacation_request_detail(" 
        "request_id, start_date, end_date, days_enjoyed, days_paid, days_total, notes" 
        ") VALUES (?, ?, ?, ?, ?, ?, ?)",
        (
            int(request_id),
            d_from,
            d_to,
            float(days_enjoyed),
            float(days_paid),
            float(total_days),
            notes,
        ),
    )

    try:
        call_proc("rrhh.sp_submit_request", (int(request_id), int(current_user.user_id)))
    except Exception as ex:
        flash(f"Se creó la solicitud pero no se pudo enviar al workflow: {ex}", "danger")
        return redirect(url_for("modulos.vacaciones"))

    flash("Solicitud enviada. Primero aprueba tu jefe directo y luego RRHH.", "success")
    return redirect(url_for("modulos.vacaciones"))


@modulos_bp.route("/vacaciones/aprobaciones")
@login_required
def vacaciones_aprobaciones():
    is_backoffice = _is_admin_or_rrhh()
    es_jefe = bool(getattr(current_user, "es_jefe", False))

    if not (is_backoffice or es_jefe):
        flash("No tienes permisos para acceder a esta sección.", "warning")
        return redirect(url_for("modulos.dashboard"))

    sql = (
        "SELECT "
        "  s.step_id, s.request_id, s.step_no, s.assigned_to_user_id, s.status AS step_status, "
        "  r.employee_id, r.created_at, r.submitted_at, "
        "  e.doc_number, (e.first_name + ' ' + e.last_name) AS employee_name, "
        "  d.start_date, d.end_date, d.days_enjoyed, d.days_paid, d.days_total, d.notes "
        "FROM rrhh.wf_request_step s "
        "JOIN rrhh.wf_request r ON r.request_id=s.request_id "
        "JOIN rrhh.vacation_request_detail d ON d.request_id=r.request_id "
        "JOIN rrhh.hr_employee e ON e.employee_id=r.employee_id "
        "WHERE r.request_type=? "
        "  AND r.status='SUBMITTED' "
        "  AND s.status='PENDING' "
        "  AND s.step_no = (" 
        "    SELECT MIN(s2.step_no) "
        "    FROM rrhh.wf_request_step s2 "
        "    WHERE s2.request_id=s.request_id AND s2.status='PENDING'" 
        "  )"
    )

    params = [REQUEST_TYPE_VACACIONES]
    if not is_backoffice:
        sql += " AND s.assigned_to_user_id=?"
        params.append(int(current_user.user_id))

    sql += " ORDER BY r.created_at DESC"

    steps = fetch_all(sql, tuple(params))

    # backoffice: histórico
    history = []
    if is_backoffice:
        history = fetch_all(
            "SELECT TOP 200 "
            "  r.request_id, r.status, r.employee_id, r.created_at, r.closed_at, "
            "  e.doc_number, (e.first_name + ' ' + e.last_name) AS employee_name, "
            "  d.start_date, d.end_date, d.days_enjoyed, d.days_paid, d.days_total "
            "FROM rrhh.wf_request r "
            "JOIN rrhh.vacation_request_detail d ON d.request_id=r.request_id "
            "JOIN rrhh.hr_employee e ON e.employee_id=r.employee_id "
            "WHERE r.request_type=? "
            "ORDER BY r.created_at DESC",
            (REQUEST_TYPE_VACACIONES,),
        )

    return render_template(
        "modulos/vacaciones_aprobaciones.html",
        steps=steps or [],
        history=history or [],
        is_backoffice=is_backoffice,
        is_admin=_is_admin(),
        is_rrhh=_is_rrhh(),
    )


@modulos_bp.route("/vacaciones/aprobaciones/accion", methods=["POST"])
@login_required
def vacaciones_aprobaciones_accion():
    step_id = int(request.form.get("step_id") or 0)
    action = (request.form.get("action") or "").strip().upper()
    comment = (request.form.get("comment") or "").strip() or None

    if step_id <= 0:
        flash("Paso inválido.", "warning")
        return redirect(url_for("modulos.vacaciones_aprobaciones"))

    if action not in ("APPROVE", "REJECT"):
        flash("Acción inválida.", "warning")
        return redirect(url_for("modulos.vacaciones_aprobaciones"))

    step = fetch_one(
        "SELECT s.step_id, s.request_id, s.step_no, s.assigned_to_user_id, s.assigned_to_role, s.status, "
        "       r.request_type "
        "FROM rrhh.wf_request_step s "
        "JOIN rrhh.wf_request r ON r.request_id=s.request_id "
        "WHERE s.step_id=?",
        (int(step_id),),
    )

    if (not step) or str(getattr(step, "request_type", "")).upper() != REQUEST_TYPE_VACACIONES:
        flash("El paso no existe o no corresponde a Vacaciones.", "warning")
        return redirect(url_for("modulos.vacaciones_aprobaciones"))

    if str(getattr(step, "status", "")).upper() != "PENDING":
        flash("El paso no existe o ya fue gestionado.", "warning")
        return redirect(url_for("modulos.vacaciones_aprobaciones"))

    # Validar que sea el paso activo (mínimo pendiente)
    row_min = fetch_one(
        "SELECT MIN(step_no) AS m FROM rrhh.wf_request_step WHERE request_id=? AND status='PENDING'",
        (int(step.request_id),),
    )
    if row_min and int(step.step_no) != int(getattr(row_min, "m", step.step_no) or step.step_no):
        flash("Este paso no es el paso activo actual para la solicitud.", "warning")
        return redirect(url_for("modulos.vacaciones_aprobaciones"))

    # Permisos
    is_admin = _is_admin()
    is_rrhh = _is_rrhh()

    if not is_admin:
        role = str(getattr(step, "assigned_to_role", "") or "").upper()

        # Rol MANAGER: solo el asignado
        if role == "MANAGER":
            if int(step.assigned_to_user_id) != int(current_user.user_id):
                flash("No tienes permiso para gestionar este paso.", "warning")
                return redirect(url_for("modulos.vacaciones_aprobaciones"))

        # Rol HR: requiere ser RRHH
        elif role == "HR":
            if not is_rrhh:
                flash("Solo RRHH (o administrador) puede gestionar este paso.", "warning")
                return redirect(url_for("modulos.vacaciones_aprobaciones"))

        # Fallback: si no viene rol, exigimos ser el asignado
        else:
            if int(step.assigned_to_user_id) != int(current_user.user_id):
                flash("No tienes permiso para gestionar este paso.", "warning")
                return redirect(url_for("modulos.vacaciones_aprobaciones"))

    # ¿Cuántos pasos pendientes hay actualmente?
    row_pending = fetch_one(
        "SELECT COUNT(1) AS c FROM rrhh.wf_request_step WHERE request_id=? AND status='PENDING'",
        (int(step.request_id),),
    )
    pending = int(getattr(row_pending, "c", 0) or 0) if row_pending else 0

    if action == "APPROVE":
        # Si este es el ÚLTIMO paso pendiente, cerramos + aplicamos saldo en 1 transacción
        if pending == 1:
            try:
                _approve_final_step_and_apply_balance(
                    step_id=int(step.step_id),
                    request_id=int(step.request_id),
                    step_no=int(step.step_no),
                    actor_user_id=int(current_user.user_id),
                    comment=comment,
                )
                flash("Solicitud aprobada y saldo actualizado.", "success")
            except Exception as ex:
                flash(f"No se pudo aprobar la solicitud: {ex}", "danger")
            return redirect(url_for("modulos.vacaciones_aprobaciones"))

        # Paso intermedio: solo aprobamos el step
        try:
            sql = (
                "BEGIN TRAN\n"
                "  UPDATE rrhh.wf_request_step\n"
                "     SET status='APPROVED', acted_at=GETDATE(), comment=?\n"
                "   WHERE step_id=? AND status='PENDING';\n"
                "  IF @@ROWCOUNT = 0\n"
                "  BEGIN\n"
                "    ROLLBACK TRAN;\n"
                "    RAISERROR('El paso no está pendiente o ya fue gestionado.',16,1);\n"
                "    RETURN;\n"
                "  END\n"
                "\n"
                "  INSERT INTO rrhh.wf_action(request_id, step_no, actor_user_id, action, comment)\n"
                "  VALUES (?, ?, ?, 'APPROVE', ?);\n"
                "COMMIT TRAN"
            )
            execute(
                sql,
                (
                    comment,
                    int(step.step_id),
                    int(step.request_id),
                    int(step.step_no),
                    int(current_user.user_id),
                    comment,
                ),
            )
            flash("Paso aprobado.", "success")
        except Exception as ex:
            flash(f"No se pudo aprobar el paso: {ex}", "danger")

        return redirect(url_for("modulos.vacaciones_aprobaciones"))

    # REJECT
    try:
        _reject_step_and_close_request(
            step_id=int(step.step_id),
            request_id=int(step.request_id),
            step_no=int(step.step_no),
            actor_user_id=int(current_user.user_id),
            comment=comment,
        )
        flash("Solicitud rechazada.", "success")
    except Exception as ex:
        flash(f"No se pudo rechazar la solicitud: {ex}", "danger")

    return redirect(url_for("modulos.vacaciones_aprobaciones"))


@modulos_bp.route("/vacaciones/cargar", methods=["GET", "POST"])
@login_required
def vacaciones_cargar():
    """Carga masiva de saldo de vacaciones desde Excel.

    Solo RRHH y Administrador.
    """

    if not _is_admin_or_rrhh():
        flash("No tienes permisos para acceder a esta sección.", "warning")
        return redirect(url_for("modulos.dashboard"))

    if request.method == "GET":
        batches = []
        if _table_exists("rrhh", "vacation_import_batch"):
            batches = fetch_all(
                "SELECT TOP 20 batch_id, file_name, uploaded_at, uploaded_by_user_id, "
                "       total_rows, matched_rows, updated_rows, error_rows "
                "FROM rrhh.vacation_import_batch ORDER BY uploaded_at DESC"
            )

        return render_template("modulos/vacaciones_carga.html", batches=batches or [])

    # POST
    if not _vacaciones_ready() or not _table_exists("rrhh", "vacation_import_batch"):
        flash("BD no está lista para Vacaciones. Ejecuta el patch primero.", "warning")
        return redirect(url_for("modulos.vacaciones_cargar"))

    f = request.files.get("file")
    if not f or not getattr(f, "filename", ""):
        flash("Debes adjuntar un archivo.", "warning")
        return redirect(url_for("modulos.vacaciones_cargar"))

    try:
        meta = save_upload(f, prefix="vacaciones")
    except Exception as ex:
        flash(f"No se pudo guardar el archivo: {ex}", "danger")
        return redirect(url_for("modulos.vacaciones_cargar"))

    checksum = None
    try:
        checksum = _checksum_file(meta["storage_path"])  # noqa
    except Exception:
        checksum = None

    batch_id = execute_scalar(
        "INSERT INTO rrhh.vacation_import_batch(file_name, checksum, uploaded_by_user_id) "
        "OUTPUT INSERTED.batch_id VALUES (?, ?, ?)",
        (meta["file_name"], checksum, int(current_user.user_id)),
    )

    if not batch_id:
        flash("No se pudo crear el batch de importación.", "danger")
        return redirect(url_for("modulos.vacaciones_cargar"))

    try:
        rows = _parse_vacaciones_excel(meta["storage_path"])
    except Exception as ex:
        execute(
            "UPDATE rrhh.vacation_import_batch SET status='FAILED', error_rows=1 WHERE batch_id=?",
            (int(batch_id),),
        )
        flash(f"No se pudo leer el Excel: {ex}", "danger")
        return redirect(url_for("modulos.vacaciones_cargar"))

total = 0
matched = 0
updated = 0
errors = 0

as_of = date.today()

# Construye un mapa de empleados por cédula "normalizada" (solo dígitos)
# para evitar fallos por formatos como 1023456789 / 1023-456-789 / espacios, etc.
emp_rows = fetch_all("SELECT employee_id, doc_number FROM rrhh.hr_employee")
emp_map: dict[str, int] = {}
dup_docs: set[str] = set()
for er in emp_rows or []:
    key = _parse_doc_number(getattr(er, "doc_number", None))
    if not key:
        continue
    if key in emp_map and emp_map[key] != int(er.employee_id):
        dup_docs.add(key)
    else:
        emp_map[key] = int(er.employee_id)

for r in rows:
    total += 1

    doc_raw = r.get("doc_number")
    doc = _parse_doc_number(doc_raw) or ""
    name = (r.get("employee_name") or "").strip() or None
    used_days = float(r.get("used_days") or 0)
    avail_days = float(r.get("available_days") or 0)

    if not doc:
        errors += 1
        execute(
            "INSERT INTO rrhh.vacation_import_row(batch_id, doc_number, employee_name, used_days, available_days, employee_id, status, error_msg) "
            "VALUES (?, ?, ?, ?, ?, NULL, 'ERROR', 'Cédula vacía o inválida en el Excel')",
            (int(batch_id), str(doc_raw or "").strip() or None, name, used_days, avail_days),
        )
        continue

    if doc in dup_docs:
        errors += 1
        execute(
            "INSERT INTO rrhh.vacation_import_row(batch_id, doc_number, employee_name, used_days, available_days, employee_id, status, error_msg) "
            "VALUES (?, ?, ?, ?, ?, NULL, 'ERROR', 'Cédula duplicada en hr_employee (revisar registros)')",
            (int(batch_id), doc, name, used_days, avail_days),
        )
        continue

    emp_id = emp_map.get(doc)

    if not emp_id:
        errors += 1
        execute(
            "INSERT INTO rrhh.vacation_import_row(batch_id, doc_number, employee_name, used_days, available_days, employee_id, status, error_msg) "
            "VALUES (?, ?, ?, ?, ?, NULL, 'ERROR', 'Empleado no existe en hr_employee (cc no encontrada)')",
            (int(batch_id), doc, name, used_days, avail_days),
        )
        continue

    matched += 1

    # Upsert balance
    merge_sql = (
        "MERGE rrhh.vacation_balance AS t "
        "USING (SELECT ? AS employee_id) AS s "
        "ON t.employee_id = s.employee_id "
        "WHEN MATCHED THEN "
        "  UPDATE SET available_days=?, used_days=?, as_of_date=?, source_batch_id=?, updated_by_user_id=?, updated_at=GETDATE() "
        "WHEN NOT MATCHED THEN "
        "  INSERT (employee_id, available_days, used_days, as_of_date, source_batch_id, updated_by_user_id) "
        "  VALUES (?, ?, ?, ?, ?, ?);"
    )

    execute(
        merge_sql,
        (
            int(emp_id),
            avail_days,
            used_days,
            as_of,
            int(batch_id),
            int(current_user.user_id),
            int(emp_id),
            avail_days,
            used_days,
            as_of,
            int(batch_id),
            int(current_user.user_id),
        ),
    )
    updated += 1

    execute(
        "INSERT INTO rrhh.vacation_import_row(batch_id, doc_number, employee_name, used_days, available_days, employee_id, status, error_msg) "
        "VALUES (?, ?, ?, ?, ?, ?, 'OK', NULL)",
        (int(batch_id), doc, name, used_days, avail_days, int(emp_id)),
    )

    execute(
        "UPDATE rrhh.vacation_import_batch "
        "SET status='APPLIED', total_rows=?, matched_rows=?, updated_rows=?, error_rows=? "
        "WHERE batch_id=?",
        (int(total), int(matched), int(updated), int(errors), int(batch_id)),
    )

    flash(
        f"Importación aplicada. Total: {total}. Matched: {matched}. Actualizados: {updated}. Errores: {errors}.",
        "success",
    )
    return redirect(url_for("modulos.vacaciones_cargar"))
