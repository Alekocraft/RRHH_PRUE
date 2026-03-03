from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from typing import Dict, List, Tuple, Optional, Set

from flask import render_template, request, redirect, url_for, flash, make_response
from flask_login import login_required, current_user
from openpyxl import load_workbook

from services.rrhh_db import fetch_all, fetch_one, execute, execute_scalar
from services.upload import save_upload
from services.report_export import build_excel, build_pdf

from .modulos import modulos_bp
from .modulos_common import (
    _require_admin,
    _month_range,
    _time_to_str,
    _parse_doc_number,
    _parse_date,
    _parse_time,
    _diff_minutes,
    _checksum_file,
    _att_import_doc_column,
)


# ============================================================
# Helpers (DB metadata, default period)
# ============================================================
_COL_CACHE: Dict[Tuple[str, str], bool] = {}

def _has_column(table_schema: str, table_name: str, column_name: str) -> bool:
    """Check column existence in SQL Server via INFORMATION_SCHEMA.COLUMNS."""
    key = (f"{table_schema}.{table_name}", column_name)
    if key in _COL_CACHE:
        return _COL_CACHE[key]
    try:
        q = (
            "SELECT 1 "
            "FROM INFORMATION_SCHEMA.COLUMNS "
            "WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ? AND COLUMN_NAME = ?"
        )
        v = execute_scalar(q, (table_schema, table_name, column_name))
        _COL_CACHE[key] = bool(v)
        return _COL_CACHE[key]
    except Exception:
        _COL_CACHE[key] = False
        return False


def _latest_attendance_month(employee_id: Optional[int] = None) -> Tuple[int, int]:
    """Return (year, month) of the latest imported attendance row (global or per employee)."""
    try:
        if employee_id is None:
            d = execute_scalar("SELECT MAX(work_date) FROM rrhh.vw_attendance_import_accum", ())
        else:
            d = execute_scalar(
                "SELECT MAX(work_date) FROM rrhh.vw_attendance_import_accum WHERE employee_id = ?",
                (employee_id,),
            )
        if d:
            if isinstance(d, datetime):
                d = d.date()
            return int(d.year), int(d.month)
    except Exception:
        pass
    today = date.today()
    return today.year, today.month

# ============================================================
# Excel parsing (Primera y última / "Employee ID" format)
# ============================================================
def _to_minutes_from_total(value) -> Optional[int]:
    """
    Many exports provide "Tiempo Total" in decimal hours (e.g. 10.2 == 10h12m).
    Convert to minutes. If it's already minutes (heuristic), keep.
    """
    if value is None:
        return None
    try:
        if isinstance(value, str):
            s = value.strip()
            if not s:
                return None
            if ":" in s and len(s.split(":")) >= 2:
                hh, mm = s.split(":")[0:2]
                return max(0, int(hh) * 60 + int(mm))
            value = float(s.replace(",", "."))
        if isinstance(value, (int, float)):
            fv = float(value)
            # Heuristic: very large numbers likely already minutes
            if fv > 72:
                return int(round(fv))
            return int(round(fv * 60))
    except Exception:
        return None
    return None


def _parse_attendance_excel(file_path: str) -> List[Dict]:
    """
    Parse an attendance Excel file and return daily consolidated rows:
    (doc_number, work_date) -> first_in(min), last_out(max), total_minutes(max)
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # Find header row (some exports have a title row first)
    header_row = 1
    for r in range(1, 7):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        joined = " ".join([f"{v}".lower() for v in vals if v is not None])
        if "fecha" in joined and ("employee" in joined or "cedula" in joined or "id" in joined):
            header_row = r
            break

    header = []
    for cell in ws[header_row]:
        header.append(f"{cell.value}".strip().lower() if cell.value is not None else "")
    header_map = {name: idx for idx, name in enumerate(header)}

    def find_col(patterns):
        for name, idx in header_map.items():
            for p in patterns:
                if p in name:
                    return idx
        return None

    col_doc = find_col(
        ["employee id", "employee", "cedula", "cédula", "documento", "doc", "identific", "cc", "dni", "id"]
    )
    col_date = find_col(["fecha", "date"])
    col_in = find_col(["primera", "entrada", "first", "ingreso", "in"])
    col_out = find_col(["última", "ultima", "salida", "last", "egreso", "out"])
    col_total = find_col(["tiempo total", "total", "duracion", "duración", "hours", "horas"])

    if col_doc is None or col_date is None:
        raise ValueError("No se encontró columna de Cédula/Documento (Employee ID) y/o Fecha en el Excel.")

    grouped: Dict[Tuple[str, date], Dict] = {}
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        doc = _parse_doc_number(r[col_doc] if col_doc < len(r) else None)
        wdate = _parse_date(r[col_date] if col_date < len(r) else None)
        if not doc or not wdate:
            continue

        fin = _parse_time(r[col_in] if (col_in is not None and col_in < len(r)) else None)
        fout = _parse_time(r[col_out] if (col_out is not None and col_out < len(r)) else None)

        mins = None
        if col_total is not None and col_total < len(r):
            mins = _to_minutes_from_total(r[col_total])

        key = (doc, wdate)
        entry = grouped.get(key)
        raw = {"row": [f"{x}" if x is not None else "" for x in r]}

        if entry is None:
            grouped[key] = {
                "doc_number": doc,
                "work_date": wdate,
                "first_in": fin,
                "last_out": fout,
                "total_minutes": mins,
                "raw": [raw],
            }
        else:
            if fin and (entry["first_in"] is None or fin < entry["first_in"]):
                entry["first_in"] = fin
            if fout and (entry["last_out"] is None or fout > entry["last_out"]):
                entry["last_out"] = fout
            if mins is not None:
                if entry["total_minutes"] is None or mins > entry["total_minutes"]:
                    entry["total_minutes"] = mins
            entry["raw"].append(raw)

    rows = []
    for v in grouped.values():
        if v["total_minutes"] is None and v["first_in"] and v["last_out"]:
            try:
                v["total_minutes"] = max(0, _diff_minutes(v["first_in"], v["last_out"]))
            except Exception:
                v["total_minutes"] = None
        v["raw_text"] = json.dumps(v["raw"], ensure_ascii=False)
        rows.append(v)

    return rows


# ============================================================
# Attendance effective view selection (accumulative)
# ============================================================
def _attendance_effective_view() -> str:
    """
    Prefer new accumulative view (built over import rows) if installed,
    otherwise fall back to old view.
    """
    v2 = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.VIEWS "
        "WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='vw_attendance_effective_accum'"
    )
    return "rrhh.vw_attendance_effective_accum" if v2 else "rrhh.vw_attendance_effective"


# ============================================================
# Cross-module helpers (vacaciones / incapacidad / chequera / feriados)
# ============================================================
def _holidays_between(d_from: date, d_to: date) -> Set[date]:
    """Obtiene festivos entre dos fechas.

    Soporta variaciones comunes del esquema (nombre de tabla/columna) para evitar
    marcar festivos como faltantes en Asistencia.
    """
    out: Set[date] = set()

    def _table_exists(schema: str, name: str) -> bool:
        try:
            return (
                fetch_one(
                    "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA=? AND TABLE_NAME=?",
                    (schema, name),
                )
                is not None
            )
        except Exception:
            return False

    def _pick_col(schema: str, table: str, candidates: List[str]) -> Optional[str]:
        for c in candidates:
            if _has_column(schema, table, c):
                return c
        return None

    # Tablas candidatas (en orden de preferencia)
    candidates = [
        ("rrhh", "hr_holiday"),
        ("rrhh", "hr_holidays"),
        ("rrhh", "holiday"),
        ("rrhh", "holidays"),
        ("rrhh", "festivo"),
        ("rrhh", "festivos"),
        ("rrhh", "holiday_calendar"),
        ("rrhh", "calendar_day"),
    ]

    date_cols = [
        "holiday_date",
        "festive_date",
        "work_date",
        "date",
        "day",
        "fecha",
        "dia",
    ]

    for schema, table in candidates:
        if not _table_exists(schema, table):
            continue

        date_col = _pick_col(schema, table, date_cols)
        if not date_col:
            continue

        wh = [f"{date_col} BETWEEN ? AND ?"]
        params = [d_from, d_to]

        # Filtros típicos de activo/flag de festivo
        if _has_column(schema, table, "is_active"):
            wh.append("is_active=1")
        elif _has_column(schema, table, "active"):
            wh.append("active=1")
        elif _has_column(schema, table, "is_holiday"):
            wh.append("is_holiday=1")

        q = f"SELECT {date_col} AS d FROM {schema}.{table} WHERE " + " AND ".join(wh)

        try:
            rows = fetch_all(q, tuple(params))
        except Exception:
            continue

        for r in rows:
            d = getattr(r, "d", None)
            if not d:
                continue
            # pyodbc puede devolver datetime; normalizamos a date
            try:
                dd = d.date() if hasattr(d, "date") else d
            except Exception:
                continue
            if isinstance(dd, date):
                out.add(dd)

    return out

def _approved_incap_days(emp_ids: List[int], d_from: date, d_to: date) -> Set[Tuple[int, date]]:
    exists = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='medical_leave_request'"
    )
    if not exists or not emp_ids:
        return set()
    rows = fetch_all(
        f"SELECT r.employee_id, d.start_date, d.end_date "
        f"FROM rrhh.wf_request r "
        f"JOIN rrhh.medical_leave_request d ON d.request_id=r.request_id "
        f"WHERE r.request_type='INCAPACIDAD' AND r.status='APPROVED' "
        f"  AND r.employee_id IN ({','.join(['?']*len(emp_ids))}) "
        f"  AND d.end_date >= ? AND d.start_date <= ?",
        tuple(emp_ids + [d_from, d_to]),
    )
    out: Set[Tuple[int, date]] = set()
    for rr in rows:
        emp = int(rr.employee_id)
        start = max(d_from, rr.start_date)
        end = min(d_to, rr.end_date)
        cur = start
        while cur <= end:
            if cur.weekday() < 5:
                out.add((emp, cur))
            cur += timedelta(days=1)
    return out


def _approved_vac_days(emp_ids: List[int], d_from: date, d_to: date, holidays: Set[date]) -> Set[Tuple[int, date]]:
    exists = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='vacation_request'"
    )
    if not exists or not emp_ids:
        return set()
    rows = fetch_all(
        f"SELECT employee_id, start_date, end_date "
        f"FROM rrhh.vacation_request "
        f"WHERE status='APPROVED' "
        f"  AND employee_id IN ({','.join(['?']*len(emp_ids))}) "
        f"  AND end_date >= ? AND start_date <= ?",
        tuple(emp_ids + [d_from, d_to]),
    )
    out: Set[Tuple[int, date]] = set()
    for rr in rows:
        emp = int(rr.employee_id)
        start = max(d_from, rr.start_date)
        end = min(d_to, rr.end_date)
        cur = start
        while cur <= end:
            # days that should not require attendance
            if cur.weekday() < 5 and (cur not in holidays):
                out.add((emp, cur))
            cur += timedelta(days=1)
    return out


def _approved_chequera_halfday(emp_ids: List[int], d_from: date, d_to: date) -> Dict[Tuple[int, date], str]:
    exists = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='timebook_request'"
    )
    if not exists or not emp_ids:
        return {}
    rows = fetch_all(
        f"SELECT employee_id, request_date, slot "
        f"FROM rrhh.timebook_request "
        f"WHERE status='APPROVED' "
        f"  AND employee_id IN ({','.join(['?']*len(emp_ids))}) "
        f"  AND request_date BETWEEN ? AND ?",
        tuple(emp_ids + [d_from, d_to]),
    )
    out: Dict[Tuple[int, date], str] = {}
    for r in rows:
        out[(int(r.employee_id), r.request_date)] = (r.slot or "AM")
    return out


# ============================================================
# UI: listado (buscable) + detalle
# ============================================================
@modulos_bp.route("/asistencia")
@login_required
def asistencia():
    if not _require_admin():
        flash("No tienes permisos para acceder a esta sección.", "warning")
        return redirect(url_for("modulos.dashboard"))

    today = date.today()
    if request.args.get("year") and request.args.get("month"):
        y = int(request.args.get("year", today.year))
        m = int(request.args.get("month", today.month))
    else:
        y, m = _latest_attendance_month(None)
    q = (request.args.get("q") or "").strip()
    dept = (request.args.get("department") or "all").strip()

    page = max(1, int(request.args.get("page", 1)))
    page_size = 25
    offset = (page - 1) * page_size

    d_from, d_to = _month_range(y, m)
    days = [d_from + timedelta(days=i) for i in range((d_to - d_from).days + 1)]
    workdays = [d for d in days if d.weekday() < 5]  # Lun-Vie

    holidays = _holidays_between(d_from, d_to)
    workdays = [d for d in workdays if d not in holidays]

    departments = fetch_all(
        "SELECT DISTINCT department AS name FROM rrhh.hr_employee "
        "WHERE is_active=1 AND department IS NOT NULL AND LTRIM(RTRIM(department))<>'' "
        "ORDER BY department"
    )
    dept_list = [d.name for d in departments]

    where = ["is_active=1"]
    params: List = []
    if dept and dept.lower() != "all":
        where.append("department = ?")
        params.append(dept)

    if q:
        if q.isdigit():
            where.append("doc_number LIKE ?")
            params.append(f"%{q}%")
        else:
            where.append("(first_name LIKE ? OR last_name LIKE ? OR (first_name + ' ' + last_name) LIKE ?)")
            like = f"%{q}%"
            params.extend([like, like, like])

    where_sql = " AND ".join(where)
    total_count = int(execute_scalar(f"SELECT COUNT(1) FROM rrhh.hr_employee WHERE {where_sql}", tuple(params)) or 0)

    empleados = fetch_all(
        f"SELECT employee_id, doc_number, first_name, last_name, department "
        f"FROM rrhh.hr_employee WHERE {where_sql} "
        f"ORDER BY last_name, first_name "
        f"OFFSET ? ROWS FETCH NEXT ? ROWS ONLY",
        tuple(params + [offset, page_size]),
    )
    emp_ids = [int(e.employee_id) for e in empleados]

    total_pages = max(1, (total_count + page_size - 1) // page_size)

    if not emp_ids:
        return render_template(
            "modulos/asistencia.html",
            year=y,
            month=m,
            q=q,
            department=dept,
            departments=dept_list,
            page=page,
            total_pages=total_pages,
            total_count=total_count,
            rows=[],
        )

    view = _attendance_effective_view()

    att_rows = fetch_all(
        f"SELECT employee_id, work_date, first_in, last_out, total_minutes, has_manual_override "
        f"FROM {view} "
        f"WHERE work_date BETWEEN ? AND ? AND employee_id IN ({','.join(['?']*len(emp_ids))})",
        tuple([d_from, d_to] + emp_ids),
    )
    att_map = {(int(r.employee_id), r.work_date): r for r in att_rows}

    # Work from home
    wfh_exists = (
        fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='wfh_day'")
        is not None
    )
    wfh_map = set()
    if wfh_exists:
        wfh_rows = fetch_all(
            f"SELECT employee_id, work_date FROM rrhh.wfh_day "
            f"WHERE is_active=1 AND work_date BETWEEN ? AND ? AND employee_id IN ({','.join(['?']*len(emp_ids))})",
            tuple([d_from, d_to] + emp_ids),
        )
        wfh_map = {(int(r.employee_id), r.work_date) for r in wfh_rows}

    # Cross modules
    incap_set = _approved_incap_days(emp_ids, d_from, d_to)
    vac_set = _approved_vac_days(emp_ids, d_from, d_to, holidays)
    cheq_map = _approved_chequera_halfday(emp_ids, d_from, d_to)

    # time reduction
    tr_exists = (
        fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='time_reduction_request'")
        is not None
    )
    reduction_map: Dict[Tuple[int, date], int] = {}
    if tr_exists:
        tr_rows = fetch_all(
            "SELECT r.employee_id, t.reduction_date, t.minutes "
            "FROM rrhh.wf_request r "
            "JOIN rrhh.time_reduction_request t ON t.request_id = r.request_id "
            "WHERE r.request_type = 'TIME_REDUCTION' AND r.status = 'APPROVED' "
            "AND t.reduction_date BETWEEN ? AND ?",
            (d_from, d_to),
        )
        for rr in tr_rows:
            reduction_map[(int(rr.employee_id), rr.reduction_date)] = int(rr.minutes or 60)

    # periodic flex rule
    flex_exists = (
        fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='time_flexible_rule'")
        is not None
    )
    if flex_exists:
        rules = fetch_all(
            "SELECT employee_id, weekday, minutes, valid_from, valid_to "
            "FROM rrhh.time_flexible_rule "
            "WHERE is_active=1 AND valid_from <= ? AND (valid_to IS NULL OR valid_to >= ?)",
            (d_to, d_from),
        )
        for rrule in rules:
            emp_id = int(rrule.employee_id)
            if emp_id not in emp_ids:
                continue
            wd = int(rrule.weekday)
            mins = int(getattr(rrule, "minutes", 60) or 60)
            vf = rrule.valid_from
            vt = rrule.valid_to or d_to
            start = max(d_from, vf)
            end = min(d_to, vt)
            cur = start
            while cur <= end:
                if (cur.weekday() + 1) == wd:
                    prev = reduction_map.get((emp_id, cur), 0) or 0
                    reduction_map[(emp_id, cur)] = max(int(prev), mins)
                cur += timedelta(days=1)

    # shifts
    shifts = fetch_all(
        f"SELECT sa.employee_id, sa.valid_from, sa.valid_to, sd.shift_code, sd.start_time, sd.end_time "
        f"FROM rrhh.shift_assignment sa "
        f"JOIN rrhh.shift_definition sd ON sd.shift_id = sa.shift_id "
        f"WHERE sa.employee_id IN ({','.join(['?']*len(emp_ids))}) "
        f"  AND sa.valid_from <= ? AND (sa.valid_to IS NULL OR sa.valid_to >= ?) "
        f"ORDER BY sa.employee_id, sa.valid_from",
        tuple(emp_ids + [d_to, d_from]),
    )
    shifts_by_emp: Dict[int, List] = {}
    for s in shifts:
        shifts_by_emp.setdefault(int(s.employee_id), []).append(s)

    def shift_for(emp_id: int, day: date):
        lst = shifts_by_emp.get(emp_id, [])
        best = None
        for s in lst:
            if s.valid_from <= day and (s.valid_to is None or s.valid_to >= day):
                best = s
        return best

    rows = []
    for e in empleados:
        emp_id = int(e.employee_id)
        summary = {"cumple": 0, "incompleto": 0, "faltante": 0, "casa": 0, "justificado": 0, "sin_turno": 0}

        for d in workdays:
            if (emp_id, d) in incap_set or (emp_id, d) in vac_set:
                summary["justificado"] += 1
                continue

            sh = shift_for(emp_id, d)
            if not sh:
                summary["sin_turno"] += 1
                continue

            required = _diff_minutes(sh.start_time, sh.end_time)

            if (emp_id, d) in cheq_map:
                required = int((required + 1) // 2)

            required -= reduction_map.get((emp_id, d), 0)
            required = max(0, required)

            att = att_map.get((emp_id, d))
            if att is None or att.total_minutes is None:
                if (emp_id, d) in wfh_map:
                    summary["casa"] += 1
                    summary["cumple"] += 1
                else:
                    summary["faltante"] += 1
                continue

            mins = int(att.total_minutes or 0)
            if mins >= required:
                summary["cumple"] += 1
            else:
                summary["incompleto"] += 1

        sh_today = shift_for(emp_id, date.today())
        shift_label = sh_today.shift_code if sh_today else "—"

        rows.append(
            {
                "employee_id": emp_id,
                "doc_number": e.doc_number,
                "name": f"{e.first_name} {e.last_name}",
                "department": getattr(e, "department", None) or "—",
                "shift": shift_label,
                "summary": summary,
            }
        )

    return render_template(
        "modulos/asistencia.html",
        year=y,
        month=m,
        q=q,
        department=dept,
        departments=dept_list,
        page=page,
        total_pages=total_pages,
        total_count=total_count,
        rows=rows,
    )


@modulos_bp.route("/asistencia/export/<fmt>")
@login_required
def asistencia_export(fmt: str):
    """Exporta el consolidado mensual de asistencia (solo RRHH/Admin)."""
    if not _require_admin():
        flash("No tienes permisos para acceder a esta sección.", "warning")
        return redirect(url_for("modulos.dashboard"))

    fmt = (fmt or "").lower().strip()
    if fmt not in ("excel", "pdf"):
        flash("Formato de exportación no soportado.", "warning")
        return redirect(url_for("modulos.asistencia"))

    today = date.today()
    if request.args.get("year") and request.args.get("month"):
        y = int(request.args.get("year", today.year))
        m = int(request.args.get("month", today.month))
    else:
        y, m = _latest_attendance_month(None)

    q = (request.args.get("q") or "").strip()
    dept = (request.args.get("department") or "all").strip()

    d_from, d_to = _month_range(y, m)
    days = [d_from + timedelta(days=i) for i in range((d_to - d_from).days + 1)]
    workdays = [d for d in days if d.weekday() < 5]  # Lun-Vie
    holidays = _holidays_between(d_from, d_to)
    workdays = [d for d in workdays if d not in holidays]

    # Empleados (sin paginación en export)
    where = ["is_active=1"]
    params: List = []
    if dept and dept.lower() != "all":
        where.append("department = ?")
        params.append(dept)

    if q:
        if q.isdigit():
            where.append("doc_number LIKE ?")
            params.append(f"%{q}%")
        else:
            where.append("(first_name LIKE ? OR last_name LIKE ? OR (first_name + ' ' + last_name) LIKE ?)")
            like = f"%{q}%"
            params.extend([like, like, like])

    where_sql = " AND ".join(where)
    empleados = fetch_all(
        f"SELECT employee_id, doc_number, first_name, last_name, department "
        f"FROM rrhh.hr_employee WHERE {where_sql} "
        f"ORDER BY last_name, first_name",
        tuple(params),
    )
    emp_ids = [int(e.employee_id) for e in empleados]

    headers = ["Cédula", "Nombre", "Departamento", "Turno", "Cumple", "Incompleto", "Faltante", "Casa", "Justificado", "Sin turno"]
    title = f"Asistencia {y:04d}-{m:02d}"

    if not emp_ids:
        rows = []
    else:
        view = _attendance_effective_view()
        att_rows = fetch_all(
            f"SELECT employee_id, work_date, first_in, last_out, total_minutes, has_manual_override "
            f"FROM {view} "
            f"WHERE work_date BETWEEN ? AND ? AND employee_id IN ({','.join(['?']*len(emp_ids))})",
            tuple([d_from, d_to] + emp_ids),
        )
        att_map = {(int(r.employee_id), r.work_date): r for r in att_rows}

        # Trabajo en casa
        wfh_exists = (
            fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='wfh_day'")
            is not None
        )
        wfh_map = set()
        if wfh_exists:
            wfh_rows = fetch_all(
                f"SELECT employee_id, work_date FROM rrhh.wfh_day "
                f"WHERE is_active=1 AND work_date BETWEEN ? AND ? AND employee_id IN ({','.join(['?']*len(emp_ids))})",
                tuple([d_from, d_to] + emp_ids),
            )
            wfh_map = {(int(r.employee_id), r.work_date) for r in wfh_rows}

        # Incapacidad / Vacaciones / Chequera
        incap_set = _approved_incap_days(emp_ids, d_from, d_to)
        vac_set = _approved_vac_days(emp_ids, d_from, d_to, holidays)
        cheq_map = _approved_chequera_halfday(emp_ids, d_from, d_to)

        # Turnos para el rango
        shifts = fetch_all(
            f"SELECT sa.employee_id, sa.valid_from, sa.valid_to, sd.shift_code, sd.start_time, sd.end_time "
            f"FROM rrhh.shift_assignment sa "
            f"JOIN rrhh.shift_definition sd ON sd.shift_id = sa.shift_id "
            f"WHERE sa.employee_id IN ({','.join(['?']*len(emp_ids))}) "
            f"AND sa.valid_from <= ? AND (sa.valid_to IS NULL OR sa.valid_to >= ?) "
            f"ORDER BY sa.employee_id, sa.valid_from",
            tuple(emp_ids + [d_to, d_from]),
        )
        shift_by_emp: Dict[int, List] = {}
        for s in shifts:
            shift_by_emp.setdefault(int(s.employee_id), []).append(s)

        def shift_for(emp_id: int, day: date):
            best = None
            for s in shift_by_emp.get(emp_id, []):
                if s.valid_from <= day and (s.valid_to is None or s.valid_to >= day):
                    best = s
            return best

        # Reducciones por hora flexible/solicitudes de reducción
        reduction_map: Dict[Tuple[int, date], int] = {}
        tr_exists = (
            fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='time_reduction_request'")
            is not None
        )
        if tr_exists:
            tr = fetch_all(
                f"SELECT r.employee_id, t.reduction_date, t.minutes "
                f"FROM rrhh.wf_request r "
                f"JOIN rrhh.time_reduction_request t ON t.request_id = r.request_id "
                f"WHERE r.request_type='TIME_REDUCTION' AND r.status='APPROVED' "
                f"AND t.reduction_date BETWEEN ? AND ? AND r.employee_id IN ({','.join(['?']*len(emp_ids))})",
                tuple([d_from, d_to] + emp_ids),
            )
            for rr in tr:
                reduction_map[(int(rr.employee_id), rr.reduction_date)] = int(rr.minutes or 60)

        flex_exists = (
            fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='time_flexible_rule'")
            is not None
        )
        if flex_exists:
            rules = fetch_all(
                f"SELECT employee_id, weekday, minutes, valid_from, valid_to "
                f"FROM rrhh.time_flexible_rule "
                f"WHERE is_active=1 AND valid_from <= ? AND (valid_to IS NULL OR valid_to >= ?) "
                f"AND employee_id IN ({','.join(['?']*len(emp_ids))})",
                tuple([d_to, d_from] + emp_ids),
            )
            for rrule in rules:
                emp_id = int(rrule.employee_id)
                wd = int(rrule.weekday)
                mins = int(getattr(rrule, "minutes", 60) or 60)
                vf = rrule.valid_from
                vt = rrule.valid_to or d_to
                start = max(d_from, vf)
                end = min(d_to, vt)
                cur = start
                while cur <= end:
                    if (cur.weekday() + 1) == wd:
                        prev = reduction_map.get((emp_id, cur), 0) or 0
                        reduction_map[(emp_id, cur)] = max(int(prev), mins)
                    cur += timedelta(days=1)

        # Construir filas
        out_rows = []
        for e in empleados:
            emp_id = int(e.employee_id)
            summary = {"cumple": 0, "incompleto": 0, "faltante": 0, "casa": 0, "justificado": 0, "sin_turno": 0}

            for d in workdays:
                if (emp_id, d) in incap_set or (emp_id, d) in vac_set:
                    summary["justificado"] += 1
                    continue

                sh = shift_for(emp_id, d)
                if not sh:
                    summary["sin_turno"] += 1
                    continue

                required = _diff_minutes(sh.start_time, sh.end_time)

                if (emp_id, d) in cheq_map:
                    required = int((required + 1) // 2)

                required -= reduction_map.get((emp_id, d), 0)
                required = max(0, required)

                att = att_map.get((emp_id, d))
                if att is None or att.total_minutes is None:
                    if (emp_id, d) in wfh_map:
                        summary["casa"] += 1
                        summary["cumple"] += 1
                    else:
                        summary["faltante"] += 1
                    continue

                mins = int(att.total_minutes or 0)
                if mins >= required:
                    summary["cumple"] += 1
                else:
                    summary["incompleto"] += 1

            sh_today = shift_for(emp_id, date.today())
            shift_label = sh_today.shift_code if sh_today else "—"

            out_rows.append([
                e.doc_number,
                f"{e.first_name} {e.last_name}",
                getattr(e, "department", None) or "—",
                shift_label,
                summary["cumple"],
                summary["incompleto"],
                summary["faltante"],
                summary["casa"],
                summary["justificado"],
                summary["sin_turno"],
            ])

        rows = out_rows

    if fmt == "excel":
        payload = build_excel(title, headers, rows)
        resp = make_response(payload)
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        resp.headers["Content-Disposition"] = f'attachment; filename="asistencia_{y:04d}-{m:02d}.xlsx"'
        return resp

    subtitle = [
        f"Periodo: {y:04d}-{m:02d}",
        f"Filtros: departamento={dept or 'all'}; búsqueda={q or '—'}",
    ]
    payload = build_pdf(title, subtitle, headers, rows)
    resp = make_response(payload)
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f'attachment; filename="asistencia_{y:04d}-{m:02d}.pdf"'
    return resp


@modulos_bp.route("/asistencia/detalle/export/<fmt>")
@login_required
def asistencia_detalle_export(fmt: str):
    """Exporta el detalle de asistencia por colaborador (solo RRHH/Admin)."""
    if not _require_admin():
        flash("No tienes permisos para acceder a esta sección.", "warning")
        return redirect(url_for("modulos.dashboard"))

    fmt = (fmt or "").lower().strip()
    if fmt not in ("excel", "pdf"):
        flash("Formato de exportación no soportado.", "warning")
        return redirect(url_for("modulos.asistencia"))

    employee_id = int(request.args.get("employee_id"))
    today = date.today()
    if request.args.get("year") and request.args.get("month"):
        y = int(request.args.get("year", today.year))
        m = int(request.args.get("month", today.month))
    else:
        y, m = _latest_attendance_month(employee_id)

    d_from, d_to = _month_range(y, m)
    emp = fetch_one(
        "SELECT employee_id, doc_number, first_name, last_name, department "
        "FROM rrhh.hr_employee WHERE employee_id=?",
        (employee_id,),
    )
    if not emp:
        flash("Empleado no encontrado.", "warning")
        return redirect(url_for("modulos.asistencia", year=y, month=m))

    # Reutilizar la lógica del detalle construyendo items
    days = [d_from + timedelta(days=i) for i in range((d_to - d_from).days + 1)]
    holidays = _holidays_between(d_from, d_to)
    workdays = [d for d in days if d.weekday() < 5 and d not in holidays]

    view = _attendance_effective_view()
    att_rows = fetch_all(
        f"SELECT employee_id, work_date, first_in, last_out, total_minutes, has_manual_override "
        f"FROM {view} WHERE employee_id=? AND work_date BETWEEN ? AND ?",
        (employee_id, d_from, d_to),
    )
    att_map = {r.work_date: r for r in att_rows}

    shifts = fetch_all(
        "SELECT sa.valid_from, sa.valid_to, sd.shift_code, sd.start_time, sd.end_time "
        "FROM rrhh.shift_assignment sa "
        "JOIN rrhh.shift_definition sd ON sd.shift_id = sa.shift_id "
        "WHERE sa.employee_id=? AND sa.valid_from <= ? AND (sa.valid_to IS NULL OR sa.valid_to >= ?) "
        "ORDER BY sa.valid_from",
        (employee_id, d_to, d_from),
    )

    def shift_for(day: date):
        best = None
        for s in shifts:
            if s.valid_from <= day and (s.valid_to is None or s.valid_to >= day):
                best = s
        return best

    # Cross modules sets for this employee
    incap_set = _approved_incap_days([employee_id], d_from, d_to)
    vac_set = _approved_vac_days([employee_id], d_from, d_to, holidays)
    cheq_map = _approved_chequera_halfday([employee_id], d_from, d_to)

    # Trabajo en casa
    wfh_exists = (
        fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='wfh_day'")
        is not None
    )
    wfh_set = set()
    if wfh_exists:
        wfh = fetch_all(
            "SELECT work_date FROM rrhh.wfh_day WHERE is_active=1 AND employee_id=? AND work_date BETWEEN ? AND ?",
            (employee_id, d_from, d_to),
        )
        wfh_set = {r.work_date for r in wfh}

    # Reduction map (time reduction + flex rule)
    reduction_map: Dict[date, int] = {}
    tr_exists = (
        fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='time_reduction_request'")
        is not None
    )
    if tr_exists:
        tr = fetch_all(
            "SELECT t.reduction_date, t.minutes "
            "FROM rrhh.wf_request r "
            "JOIN rrhh.time_reduction_request t ON t.request_id = r.request_id "
            "WHERE r.request_type='TIME_REDUCTION' AND r.status='APPROVED' "
            "AND r.employee_id=? AND t.reduction_date BETWEEN ? AND ?",
            (employee_id, d_from, d_to),
        )
        for rr in tr:
            reduction_map[rr.reduction_date] = int(rr.minutes or 60)

    flex_exists = (
        fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='time_flexible_rule'")
        is not None
    )
    if flex_exists:
        rules = fetch_all(
            "SELECT weekday, minutes, valid_from, valid_to "
            "FROM rrhh.time_flexible_rule "
            "WHERE is_active=1 AND employee_id=? AND valid_from <= ? AND (valid_to IS NULL OR valid_to >= ?)",
            (employee_id, d_to, d_from),
        )
        for rrule in rules:
            wd = int(rrule.weekday)
            mins = int(getattr(rrule, "minutes", 60) or 60)
            vf = rrule.valid_from
            vt = rrule.valid_to or d_to
            start = max(d_from, vf)
            end = min(d_to, vt)
            cur = start
            while cur <= end:
                if (cur.weekday() + 1) == wd:
                    prev = reduction_map.get(cur, 0) or 0
                    reduction_map[cur] = max(int(prev), mins)
                cur += timedelta(days=1)

    out = []
    for d in workdays:
        if (employee_id, d) in incap_set:
            out.append([d.isoformat(), "INC", "Incapacidad aprobada", "", "", ""])
            continue
        if (employee_id, d) in vac_set:
            out.append([d.isoformat(), "VAC", "Vacaciones aprobadas", "", "", ""])
            continue

        sh = shift_for(d)
        if not sh:
            out.append([d.isoformat(), "ST", "Sin turno asignado", "", "", ""])
            continue

        required = _diff_minutes(sh.start_time, sh.end_time)
        if (employee_id, d) in cheq_map:
            required = int((required + 1) // 2)
        required -= reduction_map.get(d, 0)
        required = max(0, required)

        att = att_map.get(d)
        if att is None or att.total_minutes is None:
            if d in wfh_set:
                out.append([d.isoformat(), "CASA", f"Trabajo en casa (req {required} min)", "", "", f"{required}"])
            else:
                out.append([d.isoformat(), "F", f"Faltante (req {required} min)", "", "", f"{required}"])
            continue

        mins = int(att.total_minutes or 0)
        if mins >= required:
            st = "OK*" if getattr(att, "has_manual_override", 0) else "OK"
            out.append([d.isoformat(), st, f"Cumple (req {required} min)", f"{att.first_in or ""}", f"{att.last_out or ""}", f"{mins}"])
        else:
            out.append([d.isoformat(), "I", f"Incompleto (req {required} min)", f"{att.first_in or ""}", f"{att.last_out or ""}", f"{mins}"])

    headers = ["Fecha", "Estado", "Detalle", "Primera", "Última", "Minutos"]
    title = f"Detalle de asistencia {y:04d}-{m:02d}"
    subtitle = [f"{emp.first_name} {emp.last_name} · CC {emp.doc_number} · {getattr(emp,'department',None) or '—'}"]

    if fmt == "excel":
        payload = build_excel(title, headers, out)
        resp = make_response(payload)
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        resp.headers["Content-Disposition"] = f'attachment; filename="asistencia_detalle_{emp.doc_number}_{y:04d}-{m:02d}.xlsx"'
        return resp

    payload = build_pdf(title, subtitle, headers, out)
    resp = make_response(payload)
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f'attachment; filename="asistencia_detalle_{emp.doc_number}_{y:04d}-{m:02d}.pdf"'
    return resp

@modulos_bp.route("/asistencia/detalle")
@login_required
def asistencia_detalle():
    if not _require_admin():
        flash("No tienes permisos para acceder a esta sección.", "warning")
        return redirect(url_for("modulos.dashboard"))

    employee_id = int(request.args.get("employee_id"))
    today = date.today()
    if request.args.get("year") and request.args.get("month"):
        y = int(request.args.get("year", today.year))
        m = int(request.args.get("month", today.month))
    else:
        y, m = _latest_attendance_month(employee_id)
    d_from, d_to = _month_range(y, m)

    emp = fetch_one(
        "SELECT employee_id, doc_number, first_name, last_name, department "
        "FROM rrhh.hr_employee WHERE employee_id=?",
        (employee_id,),
    )
    if not emp:
        flash("Empleado no encontrado.", "warning")
        return redirect(url_for("modulos.asistencia", year=y, month=m))

    days = [d_from + timedelta(days=i) for i in range((d_to - d_from).days + 1)]
    holidays = _holidays_between(d_from, d_to)
    workdays = [d for d in days if d.weekday() < 5 and d not in holidays]

    view = _attendance_effective_view()
    att_rows = fetch_all(
        f"SELECT employee_id, work_date, first_in, last_out, total_minutes, has_manual_override "
        f"FROM {view} WHERE employee_id=? AND work_date BETWEEN ? AND ?",
        (employee_id, d_from, d_to),
    )
    att_map = {r.work_date: r for r in att_rows}

    shifts = fetch_all(
        "SELECT sa.valid_from, sa.valid_to, sd.shift_code, sd.start_time, sd.end_time "
        "FROM rrhh.shift_assignment sa "
        "JOIN rrhh.shift_definition sd ON sd.shift_id = sa.shift_id "
        "WHERE sa.employee_id=? AND sa.valid_from <= ? AND (sa.valid_to IS NULL OR sa.valid_to >= ?) "
        "ORDER BY sa.valid_from",
        (employee_id, d_to, d_from),
    )

    def shift_for(day: date):
        best = None
        for s in shifts:
            if s.valid_from <= day and (s.valid_to is None or s.valid_to >= day):
                best = s
        return best

    # Cross modules sets for this employee
    incap_set = _approved_incap_days([employee_id], d_from, d_to)
    vac_set = _approved_vac_days([employee_id], d_from, d_to, holidays)
    cheq_map = _approved_chequera_halfday([employee_id], d_from, d_to)

    # Work from home
    wfh_exists = (
        fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='wfh_day'")
        is not None
    )
    wfh_set = set()
    if wfh_exists:
        wfh = fetch_all(
            "SELECT work_date FROM rrhh.wfh_day WHERE is_active=1 AND employee_id=? AND work_date BETWEEN ? AND ?",
            (employee_id, d_from, d_to),
        )
        wfh_set = {r.work_date for r in wfh}

    # Reduction map (time reduction + flex rule)
    reduction_map: Dict[date, int] = {}
    tr_exists = (
        fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='time_reduction_request'")
        is not None
    )
    if tr_exists:
        tr = fetch_all(
            "SELECT t.reduction_date, t.minutes "
            "FROM rrhh.wf_request r "
            "JOIN rrhh.time_reduction_request t ON t.request_id = r.request_id "
            "WHERE r.request_type='TIME_REDUCTION' AND r.status='APPROVED' "
            "AND r.employee_id=? AND t.reduction_date BETWEEN ? AND ?",
            (employee_id, d_from, d_to),
        )
        for rr in tr:
            reduction_map[rr.reduction_date] = int(rr.minutes or 60)

    flex_exists = (
        fetch_one("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='time_flexible_rule'")
        is not None
    )
    if flex_exists:
        rules = fetch_all(
            "SELECT weekday, minutes, valid_from, valid_to "
            "FROM rrhh.time_flexible_rule "
            "WHERE is_active=1 AND employee_id=? AND valid_from <= ? AND (valid_to IS NULL OR valid_to >= ?)",
            (employee_id, d_to, d_from),
        )
        for rrule in rules:
            wd = int(rrule.weekday)
            mins = int(getattr(rrule, "minutes", 60) or 60)
            vf = rrule.valid_from
            vt = rrule.valid_to or d_to
            start = max(d_from, vf)
            end = min(d_to, vt)
            cur = start
            while cur <= end:
                if (cur.weekday() + 1) == wd:
                    prev = reduction_map.get(cur, 0) or 0
                    reduction_map[cur] = max(int(prev), mins)
                cur += timedelta(days=1)

    items = []
    for d in workdays:
        if (employee_id, d) in incap_set:
            items.append({"date": d, "status": "INC", "title": "Incapacidad aprobada", "link": None})
            continue
        if (employee_id, d) in vac_set:
            items.append({"date": d, "status": "VAC", "title": "Vacaciones aprobadas", "link": None})
            continue

        sh = shift_for(d)
        if not sh:
            items.append({"date": d, "status": "ST", "title": "Sin turno asignado", "link": None})
            continue

        required = _diff_minutes(sh.start_time, sh.end_time)
        if (employee_id, d) in cheq_map:
            required = int((required + 1) // 2)

        required -= reduction_map.get(d, 0)
        required = max(0, required)

        att = att_map.get(d)
        if att is None or att.total_minutes is None:
            if d in wfh_set:
                items.append({"date": d, "status": "CASA", "title": f"Trabajo en casa (req {required} min)", "link": url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=d.isoformat())})
            else:
                items.append({"date": d, "status": "F", "title": f"Faltante (req {required} min)", "link": url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=d.isoformat())})
            continue

        mins = int(att.total_minutes or 0)
        if mins >= required:
            st = "OK*" if int(att.has_manual_override or 0) == 1 else "OK"
            items.append({"date": d, "status": st, "title": f"{mins} min (req {required}). In {_time_to_str(att.first_in)} / Out {_time_to_str(att.last_out)}", "link": url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=d.isoformat())})
        else:
            items.append({"date": d, "status": "I", "title": f"Incompleto: {mins} min (req {required}). In {_time_to_str(att.first_in)} / Out {_time_to_str(att.last_out)}", "link": url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=d.isoformat())})

    return render_template("modulos/asistencia_detalle.html", year=y, month=m, emp=emp, items=items)



@modulos_bp.route("/asistencia/ajuste", methods=["GET", "POST"])
@login_required
def asistencia_ajuste():
    """Ajuste manual de asistencia (solo RRHH/Admin).

    - Guarda sobre rrhh.att_manual_override (si existe) y queda reflejado en vw_attendance_effective_accum.
    - Permite marcar el día como "Trabajo en casa" (rrhh.wfh_day) para justificar el día sin exigir marcaciones.
    """
    if not _require_admin():
        flash("No tienes permisos para acceder a esta sección.", "warning")
        return redirect(url_for("modulos.dashboard"))

    employee_id = int(request.args.get("employee_id") or request.form.get("employee_id") or 0)
    work_date = request.args.get("work_date") or request.form.get("work_date") or ""
    if not employee_id or not work_date:
        flash("Parámetros incompletos para ajustar asistencia.", "warning")
        return redirect(url_for("modulos.asistencia"))

    # Identidad del usuario actual para auditoría
    decided_by_user = getattr(current_user, "username", None) or getattr(current_user, "ad_username", None) or ""
    try:
        decided_by_uid = int(getattr(current_user, "user_id", 0) or 0)
    except Exception:
        decided_by_uid = 0
    if not decided_by_uid:
        try:
            decided_by_uid = int(getattr(current_user, "id", 0) or 0)
        except Exception:
            decided_by_uid = 0
    if not decided_by_uid and decided_by_user:
        # Fallback: resolver en DB
        decided_by_uid = int(
            execute_scalar(
                "SELECT TOP 1 user_id FROM rrhh.auth_user WHERE username=? OR ad_username=?",
                (decided_by_user, decided_by_user),
            )
            or 0
        )
    if not decided_by_uid:
        flash("No se pudo determinar el usuario actual para auditoría. Contacta al administrador.", "warning")
        return redirect(url_for("modulos.asistencia_detalle", employee_id=employee_id))

    # Datos efectivos (import + override)
    eff = fetch_one(
        """
        SELECT employee_id, work_date, first_in, last_out, total_minutes, has_manual_override
        FROM rrhh.vw_attendance_effective_accum
        WHERE employee_id = ? AND work_date = ?
        """,
        (employee_id, work_date),
    )

    # Override activo (si existe tabla)
    override = None
    if _has_column("rrhh", "att_manual_override", "employee_id") and _has_column("rrhh", "att_manual_override", "work_date"):
        cols = "first_in, last_out, total_minutes, is_active"
        if _has_column("rrhh", "att_manual_override", "reason"):
            cols += ", reason"
        override = fetch_one(
            f"""
            SELECT TOP 1 {cols}
            FROM rrhh.att_manual_override
            WHERE employee_id = ? AND work_date = ? AND is_active = 1
            """,
            (employee_id, work_date),
        )

    # Trabajo en casa (si existe tabla)
    wfh_exists = _has_column("rrhh", "wfh_day", "employee_id") and _has_column("rrhh", "wfh_day", "work_date")
    wfh_active = False
    if wfh_exists:
        wfh_active = (
            fetch_one(
                "SELECT 1 FROM rrhh.wfh_day WHERE employee_id=? AND work_date=? AND is_active=1",
                (employee_id, work_date),
            )
            is not None
        )

    def _touch_wfh(is_wfh: bool):
        """Activa/desactiva rrhh.wfh_day si existe."""
        if not wfh_exists:
            return

        exists = fetch_one("SELECT 1 FROM rrhh.wfh_day WHERE employee_id=? AND work_date=?", (employee_id, work_date)) is not None
        if is_wfh:
            if exists:
                q = "UPDATE rrhh.wfh_day SET is_active=1"
                params = []
                if _has_column("rrhh", "wfh_day", "updated_at"):
                    q += ", updated_at=GETDATE()"
                if _has_column("rrhh", "wfh_day", "updated_by_user_id"):
                    q += ", updated_by_user_id=?"; params.append(decided_by_uid)
                elif _has_column("rrhh", "wfh_day", "updated_by"):
                    q += ", updated_by=?"; params.append(decided_by_uid)
                q += " WHERE employee_id=? AND work_date=?"
                params.extend([employee_id, work_date])
                execute(q, tuple(params))
            else:
                cols = ["employee_id", "work_date", "is_active"]
                vals = ["?", "?", "1"]
                params = [employee_id, work_date]
                if _has_column("rrhh", "wfh_day", "created_at"):
                    cols.append("created_at"); vals.append("GETDATE()")
                if _has_column("rrhh", "wfh_day", "updated_at"):
                    cols.append("updated_at"); vals.append("GETDATE()")
                if _has_column("rrhh", "wfh_day", "created_by_user_id"):
                    cols.append("created_by_user_id"); vals.append("?"); params.append(decided_by_uid)
                elif _has_column("rrhh", "wfh_day", "created_by"):
                    cols.append("created_by"); vals.append("?"); params.append(decided_by_uid)
                if _has_column("rrhh", "wfh_day", "updated_by_user_id"):
                    cols.append("updated_by_user_id"); vals.append("?"); params.append(decided_by_uid)
                elif _has_column("rrhh", "wfh_day", "updated_by"):
                    cols.append("updated_by"); vals.append("?"); params.append(decided_by_uid)
                q = f"INSERT INTO rrhh.wfh_day ({', '.join(cols)}) VALUES ({', '.join(vals)})"
                execute(q, tuple(params))
        else:
            if exists:
                q = "UPDATE rrhh.wfh_day SET is_active=0"
                params = []
                if _has_column("rrhh", "wfh_day", "updated_at"):
                    q += ", updated_at=GETDATE()"
                if _has_column("rrhh", "wfh_day", "updated_by_user_id"):
                    q += ", updated_by_user_id=?"; params.append(decided_by_uid)
                elif _has_column("rrhh", "wfh_day", "updated_by"):
                    q += ", updated_by=?"; params.append(decided_by_uid)
                q += " WHERE employee_id=? AND work_date=?"
                params.extend([employee_id, work_date])
                execute(q, tuple(params))

    if request.method == "POST":
        action = (request.form.get("action") or "save").lower()
        is_wfh = (request.form.get("is_wfh") or "").lower() in ("1", "true", "on", "yes", "si")

        # Siempre sincroniza wfh_day con el checkbox
        _touch_wfh(is_wfh)

        if action == "disable":
            if override:
                execute(
                    """
                    UPDATE rrhh.att_manual_override
                    SET is_active = 0
                    WHERE employee_id = ? AND work_date = ? AND is_active = 1
                    """,
                    (employee_id, work_date),
                )
                flash("Ajuste manual desactivado.", "success")
            else:
                flash("No hay ajuste activo para desactivar.", "info")
            return redirect(url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=work_date))

        # Save / upsert del override manual (opcional si marcaste Trabajo en casa)
        reason = (request.form.get("reason") or "").strip()
        if not reason:
            reason = "Trabajo en casa" if is_wfh else "Ajuste manual"

        first_in = _parse_time(request.form.get("first_in"))
        last_out = _parse_time(request.form.get("last_out"))
        total_minutes = request.form.get("total_minutes")

        mins = None
        try:
            if total_minutes is not None and f"{total_minutes}".strip() != "":
                mins = int(float(f"{total_minutes}".strip()))
        except Exception:
            mins = None

        if mins is None and first_in and last_out:
            mins = _diff_minutes(first_in, last_out)

        # Si NO es Trabajo en casa, exigimos horas/minutos válidos
        if not is_wfh:
            if first_in is None or last_out is None or mins is None:
                flash("Debes indicar Hora de entrada, Hora de salida o Minutos totales válidos.", "warning")
                return redirect(url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=work_date))

        # Si es Trabajo en casa y NO hay horas/minutos, no tocamos override: solo queda wfh_day.
        if is_wfh and (first_in is None or last_out is None or mins is None):
            flash("Día marcado como Trabajo en casa.", "success")
            return redirect(url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=work_date))

        # Upsert att_manual_override con auditoría por user_id si aplica.
        if override:
            q = """UPDATE rrhh.att_manual_override
                   SET first_in = ?, last_out = ?, total_minutes = ?, is_active = 1
                """
            params = [first_in, last_out, mins]

            if _has_column("rrhh", "att_manual_override", "reason"):
                q += ", reason = ?"
                params.append(reason)

            if _has_column("rrhh", "att_manual_override", "updated_at"):
                q += ", updated_at = GETDATE()"
            if _has_column("rrhh", "att_manual_override", "updated_by_user_id"):
                q += ", updated_by_user_id = ?"
                params.append(decided_by_uid)
            elif _has_column("rrhh", "att_manual_override", "updated_by"):
                q += ", updated_by = ?"
                params.append(decided_by_uid)

            q += " WHERE employee_id = ? AND work_date = ? AND is_active = 1"
            params.extend([employee_id, work_date])
            execute(q, tuple(params))
        else:
            base_cols = ["employee_id", "work_date", "first_in", "last_out", "total_minutes", "is_active"]
            base_vals = ["?", "?", "?", "?", "?", "1"]
            params = [employee_id, work_date, first_in, last_out, mins]

            if _has_column("rrhh", "att_manual_override", "reason"):
                base_cols.append("reason"); base_vals.append("?"); params.append(reason)

            if _has_column("rrhh", "att_manual_override", "created_at"):
                base_cols.append("created_at"); base_vals.append("GETDATE()")
            if _has_column("rrhh", "att_manual_override", "updated_at"):
                base_cols.append("updated_at"); base_vals.append("GETDATE()")

            # IMPORTANT: algunas BD exigen created_by_user_id NOT NULL
            if _has_column("rrhh", "att_manual_override", "created_by_user_id"):
                base_cols.append("created_by_user_id"); base_vals.append("?"); params.append(decided_by_uid)
            elif _has_column("rrhh", "att_manual_override", "created_by"):
                base_cols.append("created_by"); base_vals.append("?"); params.append(decided_by_uid)

            if _has_column("rrhh", "att_manual_override", "updated_by_user_id"):
                base_cols.append("updated_by_user_id"); base_vals.append("?"); params.append(decided_by_uid)
            elif _has_column("rrhh", "att_manual_override", "updated_by"):
                base_cols.append("updated_by"); base_vals.append("?"); params.append(decided_by_uid)

            q = f"INSERT INTO rrhh.att_manual_override ({', '.join(base_cols)}) VALUES ({', '.join(base_vals)})"
            execute(q, tuple(params))

        flash("Ajuste manual guardado.", "success")
        return redirect(url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=work_date))

    return render_template(
        "modulos/asistencia_ajuste.html",
        eff=eff,
        employee_id=employee_id,
        work_date=work_date,
        override=override,
        wfh_active=wfh_active,
        wfh_exists=wfh_exists,
    )


# ============================================================
# Import (RRHH/Admin) - accumulative (no replace)
# ============================================================
@modulos_bp.route("/asistencia/cargar", methods=["GET", "POST"])
@login_required
def asistencia_cargar():
    if not _require_admin():
        flash("No tienes permisos para acceder a esta sección.", "warning")
        return redirect(url_for("modulos.dashboard"))

    today = date.today()
    year_no = int(request.values.get("year_no", today.year))
    month_no = int(request.values.get("month_no", today.month))

    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename:
            flash("Debes seleccionar un archivo.", "warning")
            return render_template("modulos/asistencia_cargar.html", year_no=year_no, month_no=month_no)

        try:
            meta = save_upload(f, prefix="asistencia")
            checksum = _checksum_file(meta["storage_path"])

            file_id = int(
                execute_scalar(
                    "INSERT INTO rrhh.sys_attachment(file_name, mime_type, size_bytes, storage_path, uploaded_by_user_id) "
                    "OUTPUT INSERTED.file_id "
                    "VALUES (?, ?, ?, ?, ?)",
                    (
                        meta["file_name"],
                        f.mimetype or "application/octet-stream",
                        meta["size_bytes"],
                        meta["storage_path"],
                        current_user.user_id,
                    ),
                )
            )

            batch_id = int(
                execute_scalar(
                    "INSERT INTO rrhh.att_import_batch(year_no, month_no, file_name, file_id, checksum, uploaded_by_user_id) "
                    "OUTPUT INSERTED.batch_id "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (year_no, month_no, meta["file_name"], file_id, checksum, current_user.user_id),
                )
            )

            parsed = _parse_attendance_excel(meta["storage_path"])
            doc_col = _att_import_doc_column()

            total = 0
            matched = 0
            errors = 0

            for r in parsed:
                total += 1
                emp = fetch_one(
                    "SELECT employee_id FROM rrhh.hr_employee WHERE doc_number = ? AND is_active=1",
                    (r["doc_number"],),
                )

                if not emp:
                    errors += 1
                    execute(
                        f"INSERT INTO rrhh.att_import_row(batch_id, {doc_col}, work_date, first_in, last_out, total_minutes, raw_text, load_status, error_message) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, 'ERROR', ?)",
                        (
                            batch_id,
                            r["doc_number"],
                            r["work_date"],
                            r["first_in"],
                            r["last_out"],
                            r["total_minutes"],
                            r["raw_text"],
                            "Empleado no existe (doc_number no encontrado)",
                        ),
                    )
                else:
                    matched += 1
                    execute(
                        f"INSERT INTO rrhh.att_import_row(batch_id, {doc_col}, work_date, first_in, last_out, total_minutes, raw_text, load_status, error_message) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, 'OK', NULL)",
                        (
                            batch_id,
                            r["doc_number"],
                            r["work_date"],
                            r["first_in"],
                            r["last_out"],
                            r["total_minutes"],
                            r["raw_text"],
                        ),
                    )

            flash(f"Importación acumulativa aplicada. Batch #{batch_id}. Total: {total}. Matched: {matched}. Errores: {errors}.", "success")
            return redirect(url_for("modulos.asistencia", year=year_no, month=month_no))

        except Exception:
            flash("No se pudo cargar el archivo de asistencia. Verifica el formato e intenta de nuevo.", "danger")

    return render_template("modulos/asistencia_cargar.html", year_no=year_no, month_no=month_no)
