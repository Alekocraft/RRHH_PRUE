import calendar
import hashlib
import json
import os
from datetime import date, datetime, timedelta, time
from typing import Dict, List, Tuple, Optional

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from openpyxl import load_workbook

from services.rrhh_db import fetch_all, fetch_one, execute, execute_scalar, call_proc
from services.upload import save_upload
from services.hr_employee_service import employee_can_work_from_home, get_manager_for_employee, manager_has_subordinates
from services.rrhh_security import ROLE_ADMIN, ROLE_RRHH
from services import turnos_roster_service as roster
from services import turnos_event_service as planner

modulos_bp = Blueprint("modulos", __name__)

# -------------------------
# Helpers
# -------------------------
def _require_admin():
    roles = getattr(current_user, 'roles', None) or []
    if (
        getattr(current_user, 'is_admin', False)
        or (ROLE_ADMIN in roles)
        or (ROLE_RRHH in roles)
        or ('ADMINISTRADOR' in roles)
        or ('RRHH' in roles)
    ):
        return True
    flash('No tienes permisos para acceder a esta sección.', 'warning')
    return False

def _month_range(y: int, m: int) -> Tuple[date, date]:
    first = date(y, m, 1)
    last = date(y, m, calendar.monthrange(y, m)[1])
    return first, last

def _time_to_str(t: Optional[time]) -> str:
    if t is None:
        return ""
    return t.strftime("%H:%M")

def _parse_doc_number(v) -> Optional[str]:
    if v is None:
        return None
    s = (f"{v}").strip()
    if not s:
        return None
    # solo dígitos
    digits = "".join(ch for ch in s if ch.isdigit())
    return digits or None

def _parse_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = (f"{v}").strip()
    if not s:
        return None
    # intenta yyyy-mm-dd o dd/mm/yyyy
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def _parse_time(v) -> Optional[time]:
    if v is None:
        return None
    if isinstance(v, time):
        return v.replace(microsecond=0)
    if isinstance(v, datetime):
        return v.time().replace(microsecond=0)
    s = (f"{v}").strip()
    if not s:
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except Exception:
            pass
    return None

def _diff_minutes(t1: time, t2: time) -> int:
    dt1 = datetime.combine(date(2000, 1, 1), t1)
    dt2 = datetime.combine(date(2000, 1, 1), t2)
    return int((dt2 - dt1).total_seconds() // 60)

def _checksum_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def _att_import_doc_column() -> str:
    # soporta DB vieja (employee_code) o nueva (doc_number)
    col = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS "
        "WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='att_import_row' AND COLUMN_NAME='doc_number'"
    )
    return "doc_number" if col else "employee_code"

# -------------------------
# Dashboard
# -------------------------
@modulos_bp.route("/dashboard")
@login_required
def dashboard():
    puede_trabajo_casa = employee_can_work_from_home(getattr(current_user, "employee_id", None))
    return render_template("dashboard.html", puede_trabajo_casa=puede_trabajo_casa)


@modulos_bp.route("/trabajo-casa/solicitar", methods=["GET", "POST"])
@login_required
def trabajo_casa_solicitar():
    """Solicitud simple de trabajo en casa (por empleado).

    - Visible solo si hr_employee.can_work_from_home = 1
    - Registra en rrhh.wfh_day (si existe)
    """

    employee_id = getattr(current_user, "employee_id", None)
    if employee_id is None:
        flash("Tu usuario no está asociado a un empleado. RRHH debe asignar el perfil.", "warning")
        return redirect(url_for("perfil_pendiente"))

    if not employee_can_work_from_home(employee_id):
        flash("No tienes habilitado el módulo de trabajo en casa.", "warning")
        return redirect(url_for("modulos.dashboard"))

    exists = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='wfh_day'"
    ) is not None
    if not exists:
        flash("Tabla rrhh.wfh_day no existe. Debe crearse por script/migración.", "warning")
        return render_template("modulos/trabajo_casa_solicitar.html", registros=[])

    if request.method == "POST":
        try:
            work_date = datetime.strptime(request.form.get("work_date"), "%Y-%m-%d").date()
            reason = (request.form.get("reason") or "Trabajo en casa").strip()

            created_by = getattr(current_user, "user_db_id", None)
            if created_by is None:
                created_by = getattr(current_user, "user_id", 0)

            execute(
                "MERGE rrhh.wfh_day AS t "
                "USING (SELECT ? AS employee_id, ? AS work_date) AS s "
                "ON (t.employee_id = s.employee_id AND t.work_date = s.work_date) "
                "WHEN MATCHED THEN UPDATE SET is_active=1, reason=? "
                "WHEN NOT MATCHED THEN INSERT (employee_id, work_date, reason, created_by_user_id, is_active) "
                "VALUES (?, ?, ?, ?, 1);",
                (int(employee_id), work_date, reason, int(employee_id), work_date, reason, int(created_by)),
            )
            flash("Solicitud registrada.", "success")
            return redirect(url_for("modulos.trabajo_casa_solicitar"))
        except Exception as ex:
            flash(f"No se pudo registrar la solicitud: {ex}", "danger")

    registros = fetch_all(
        "SELECT work_date, reason "
        "FROM rrhh.wfh_day "
        "WHERE employee_id=? AND is_active=1 "
        "ORDER BY work_date DESC",
        (int(employee_id),),
    )
    return render_template("modulos/trabajo_casa_solicitar.html", registros=registros)

# -------------------------
# Turnos
#  - Asignación base (turno fijo / vigente)
#  - Planificación semanal (estilo Teams: horas vs días)
# -------------------------

def _has_col(schema: str, table: str, column: str) -> bool:
    r = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_SCHEMA=? AND TABLE_NAME=? AND COLUMN_NAME=?",
        (schema, table, column),
    )
    return r is not None


@modulos_bp.route("/turnos")
@login_required
def turnos():
    """Asignación base.

    - Filtra el catálogo de turnos por grupo (CABINA / AJUSTADOR).
    - Si rrhh.hr_employee tiene la columna shift_group, también filtra empleados.
    """

    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    group = roster.normalize_group(request.args.get("group") or "CABINA")
    q = (request.args.get("q") or "").strip()

    today = date.today()

    has_emp_group = _has_col("rrhh", "hr_employee", "shift_group")

    # Employees
    where = ["e.is_active = 1"]
    params = [today, today]

    if has_emp_group:
        where.append("e.shift_group = ?")
        params.append(group)

    if q:
        where.append("(e.doc_number LIKE ? OR e.first_name LIKE ? OR e.last_name LIKE ?)")
        like = f"%{q}%"
        params.extend([like, like, like])

    rows = fetch_all(
        "SELECT e.employee_id, e.doc_number, e.first_name, e.last_name, "
        "sd.shift_code, sd.start_time, sd.end_time "
        "FROM rrhh.hr_employee e "
        "OUTER APPLY ("
        "  SELECT TOP (1) sa.shift_id "
        "  FROM rrhh.shift_assignment sa "
        "  WHERE sa.employee_id = e.employee_id "
        "    AND sa.valid_from <= ? "
        "    AND (sa.valid_to IS NULL OR sa.valid_to >= ?) "
        "  ORDER BY sa.valid_from DESC"
        ") x "
        "LEFT JOIN rrhh.shift_definition sd ON sd.shift_id = x.shift_id "
        f"WHERE {' AND '.join(where)} "
        "ORDER BY e.last_name, e.first_name",
        tuple(params),
    )

    # Shift catalog (by group)
    shifts = fetch_all(
        "SELECT shift_id, shift_code, start_time, end_time, shift_group "
        "FROM rrhh.shift_definition "
        "WHERE is_active=1 AND shift_group=? "
        "ORDER BY start_time, end_time, shift_code",
        (group,),
    )

    if group == "AJUSTADOR" and not shifts:
        flash(
            "No hay turnos con shift_group='AJUSTADOR' en rrhh.shift_definition. "
            "Tu BD actualmente tiene CABINA/ADMIN. Aplica el script sql/patch_shift_definition_ajustador.sql.",
            "warning",
        )

    return render_template(
        "modulos/turnos.html",
        empleados=rows or [],
        shifts=shifts or [],
        group=group,
        q=q,
    )


@modulos_bp.route("/turnos/asignar", methods=["POST"])
@login_required
def turnos_asignar():
    """Asigna el turno base (vigente) a un empleado.

    Importante: el template puede enviar shift_id vacío si el usuario no selecciona.
    Esto NO debe reventar con ValueError.
    """

    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    group = roster.normalize_group(request.form.get("group") or request.args.get("group") or "CABINA")

    employee_raw = (request.form.get("employee_id") or "").strip()
    shift_raw = (request.form.get("shift_id") or "").strip()

    if not employee_raw.isdigit():
        flash("Empleado inválido.", "warning")
        return redirect(url_for("modulos.turnos", group=group))

    if not shift_raw:
        flash("Debes seleccionar un horario antes de asignar.", "warning")
        return redirect(url_for("modulos.turnos", group=group))

    try:
        employee_id = int(employee_raw)
        shift_id = int(shift_raw)
    except Exception:
        flash("Empleado/turno inválido.", "warning")
        return redirect(url_for("modulos.turnos", group=group))

    valid_from_raw = (request.form.get("valid_from") or "").strip()
    valid_to_raw = (request.form.get("valid_to") or "").strip() or None

    # Si no mandan fecha, usar hoy (evita romper el flujo simple del UI)
    if not valid_from_raw:
        vf = date.today()
    else:
        vf = datetime.strptime(valid_from_raw, "%Y-%m-%d").date()

    vt = datetime.strptime(valid_to_raw, "%Y-%m-%d").date() if valid_to_raw else None

    try:
        call_proc(
            "rrhh.sp_set_shift_assignment",
            [employee_id, shift_id, vf, vt, current_user.user_id, "Asignación manual"],
        )
        flash("Turno asignado correctamente.", "success")
    except Exception as ex:
        flash(f"No se pudo asignar el turno: {ex}", "danger")

    return redirect(url_for("modulos.turnos", group=group))


@modulos_bp.route("/turnos/mes/<group>")
@login_required
def turnos_mes(group):
    """Compatibilidad con enlaces viejos (Mes · Cabina/Ajustador).

    En esta implementación lo redirigimos a la planificación semanal.
    """

    g = roster.normalize_group(group)
    return redirect(url_for("modulos.turnos_planificacion", group=g))




@modulos_bp.route("/turnos/planificacion", methods=["GET"])
@login_required
def turnos_planificacion():
    """Planificación semanal por EVENTOS (drag/resize, estilo Teams).

    - Un empleado a la vez (lista izquierda). También se puede arrastrar desde la lista.
    - Persistencia: rrhh.shift_roster_event (por horas) y rrhh.shift_roster_week (estado/política).
    """

    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    # Schema requerido
    try:
        planner.assert_schema()
    except Exception as ex:
        flash(
            f"Planner de turnos no está habilitado: {ex}. Ejecuta sql/patch_turnos_planner_events.sql.",
            "warning",
        )
        return redirect(url_for("modulos.turnos"))

    group = roster.normalize_group(request.args.get("group") or "CABINA")

    start_raw = (request.args.get("start") or "").strip()
    ws = roster.week_start(date.today())
    if start_raw:
        try:
            ws = roster.week_start(roster.parse_date(start_raw))
        except Exception:
            ws = roster.week_start(date.today())

    week_id = planner.ensure_week(ws, group, getattr(current_user, "user_id", None))
    hdr = planner.get_week_header(week_id)

    # Empleados
    employees = roster.list_employees_for_group(group) or []

    # Selección
    emp_raw = (request.args.get("emp") or "").strip()
    selected_emp_id = None
    if emp_raw.isdigit():
        selected_emp_id = int(emp_raw)
    elif employees:
        selected_emp_id = int(employees[0].employee_id)

    policy = planner.get_policy_json(week_id) or {}

    leave_blocks = []
    if selected_emp_id:
        leave_blocks = planner.list_blocking_leave_for_week(selected_emp_id, ws)

    return render_template(
        "modulos/turnos_planificacion.html",
        week_id=week_id,
        header=hdr,
        group=group,
        week_start=str(ws),
        prev_start=str(ws - timedelta(days=7)),
        next_start=str(ws + timedelta(days=7)),
        employees=employees,
        selected_emp_id=selected_emp_id,
        policy=policy,
        leave_blocks=leave_blocks,
    )


# -------------------------
# API: Planner Turnos (Eventos)
# -------------------------

@modulos_bp.route("/api/turnos/events", methods=["GET"])
@login_required
def api_turnos_events_list():
    if not _require_admin():
        return jsonify({"ok": False, "error": "forbidden"}), 403

    week_id = int(request.args.get("week_id") or 0)
    emp_id = request.args.get("emp")
    employee_id = int(emp_id) if emp_id and str(emp_id).isdigit() else None

    events = planner.list_events(week_id, employee_id)

    # Bloqueos: chequera/cumple (solo para el empleado actual)
    leave = []
    hdr = planner.get_week_header(week_id)
    if hdr and employee_id:
        leave = planner.list_blocking_leave_for_week(employee_id, hdr.week_start)

    return jsonify({"ok": True, "events": events, "leave": leave})


@modulos_bp.route("/api/turnos/events", methods=["POST"])
@login_required
def api_turnos_events_create():
    if not _require_admin():
        return jsonify({"ok": False, "error": "forbidden"}), 403

    data = request.get_json(silent=True) or {}
    try:
        week_id = int(data.get("week_id"))
        employee_id = int(data.get("employee_id"))
        group = roster.normalize_group(data.get("group") or "CABINA")
        start_dt = datetime.fromisoformat(str(data.get("start")))
        end_dt = datetime.fromisoformat(str(data.get("end")))
        event_type = (data.get("event_type") or "WORK").upper()
        leave_code = data.get("leave_code")
        notes = data.get("notes")

        event_id = planner.create_event(
            week_id,
            employee_id,
            group,
            start_dt,
            end_dt,
            event_type,
            leave_code,
            notes,
            getattr(current_user, "user_id", None),
        )
        return jsonify({"ok": True, "event_id": event_id})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 400


@modulos_bp.route("/api/turnos/events/<int:event_id>", methods=["PUT"])
@login_required
def api_turnos_events_update(event_id: int):
    if not _require_admin():
        return jsonify({"ok": False, "error": "forbidden"}), 403

    data = request.get_json(silent=True) or {}
    try:
        start_dt = datetime.fromisoformat(str(data.get("start")))
        end_dt = datetime.fromisoformat(str(data.get("end")))
        notes = data.get("notes")

        planner.update_event(event_id, start_dt, end_dt, getattr(current_user, "user_id", None), notes)
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 400


@modulos_bp.route("/api/turnos/events/<int:event_id>", methods=["DELETE"])
@login_required
def api_turnos_events_delete(event_id: int):
    if not _require_admin():
        return jsonify({"ok": False, "error": "forbidden"}), 403

    try:
        planner.delete_event(event_id)
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 400


@modulos_bp.route("/api/turnos/week/validate", methods=["GET"])
@login_required
def api_turnos_week_validate():
    if not _require_admin():
        return jsonify({"ok": False, "error": "forbidden"}), 403

    try:
        week_id = int(request.args.get("week_id") or 0)
        res = planner.validate_week(week_id)
        return jsonify({"ok": True, "result": res})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 400


@modulos_bp.route("/api/turnos/week/policy", methods=["POST"])
@login_required
def api_turnos_week_policy():
    if not _require_admin():
        return jsonify({"ok": False, "error": "forbidden"}), 403

    data = request.get_json(silent=True) or {}
    try:
        week_id = int(data.get("week_id"))

        # Política simple (opcional)
        # Si min está vacío o 0 => no se guarda política.
        min_raw = (str(data.get("min")) if data.get("min") is not None else "").strip()
        start = (str(data.get("start")) if data.get("start") else "").strip() or "00:00"
        end = (str(data.get("end")) if data.get("end") else "").strip() or "24:00"
        step_min = int(data.get("step_min") or 30)

        if not min_raw:
            planner.set_policy_json(week_id, None)
            return jsonify({"ok": True, "policy": None})

        min_val = int(min_raw)
        if min_val <= 0:
            planner.set_policy_json(week_id, None)
            return jsonify({"ok": True, "policy": None})

        pol = {"rule": {"min": min_val, "start": start, "end": end, "step_min": step_min}}
        planner.set_policy_json(week_id, pol)
        return jsonify({"ok": True, "policy": pol})

    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 400


@modulos_bp.route("/api/turnos/week/draft", methods=["POST"])
@login_required
def api_turnos_week_draft():
    if not _require_admin():
        return jsonify({"ok": False, "error": "forbidden"}), 403

    data = request.get_json(silent=True) or {}
    try:
        week_id = int(data.get("week_id"))
        # Solo asegurar DRAFT (si estaba publicado)
        planner.set_week_status_draft(week_id)
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 400


@modulos_bp.route("/api/turnos/week/publish", methods=["POST"])
@login_required
def api_turnos_week_publish():
    if not _require_admin():
        return jsonify({"ok": False, "error": "forbidden"}), 403

    data = request.get_json(silent=True) or {}
    try:
        week_id = int(data.get("week_id"))
        force = bool(data.get("force"))

        res = planner.validate_week(week_id)
        issues = res.get("issues") or []

        if issues and not force:
            return jsonify({"ok": False, "needs_confirm": True, "result": res}), 200

        planner.publish_week(week_id, getattr(current_user, "user_id", None))
        return jsonify({"ok": True, "result": res})

    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 400


@modulos_bp.route("/turnos/planificacion-legacy", methods=["GET", "POST"])
@login_required
def turnos_planificacion_legacy():
    """Planificación semanal tipo Teams (horas vs días) para CABINA / AJUSTADOR.

    - GET: muestra la semana, permite seleccionar empleado.
    - POST: guarda cambios de la semana para el empleado seleccionado.

    Persistencia:
      rrhh.shift_roster_week (cabecera)
      rrhh.shift_roster_day  (detalle por empleado + día)
    """

    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    # Asegura tablas
    has_week = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='shift_roster_week'"
    )
    has_day = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='shift_roster_day'"
    )
    if not has_week or not has_day:
        flash(
            "No existen las tablas rrhh.shift_roster_week / rrhh.shift_roster_day. Ejecuta el script rrhh.sql (sección de turnos).",
            "warning",
        )
        return redirect(url_for("modulos.turnos"))

    group = roster.normalize_group(request.values.get("group") or request.args.get("group") or "CABINA")

    start_raw = (request.values.get("start") or request.args.get("start") or "").strip()
    ws = roster.week_start(date.today())
    if start_raw:
        try:
            ws = roster.parse_date(start_raw)
            ws = roster.week_start(ws)
        except Exception:
            ws = roster.week_start(date.today())

    week_id = roster.ensure_week_row(ws, group, getattr(current_user, "user_id", None))

    # Data
    employees = roster.list_employees_for_group(group) or []

    # Shift defs for this board
    shift_defs = roster.list_shift_defs(group) or []

    # Build shift maps
    by_id = {s.shift_id: s for s in shift_defs}
    by_code = {s.shift_code: s for s in shift_defs}

    # Days in week
    days = [ws + timedelta(days=i) for i in range(7)]

    # Load roster days
    rows = roster.fetch_roster_days(week_id) or []

    matrix: Dict[int, Dict[str, str]] = {}

    def _row_value(r):
        code = (getattr(r, "code", None) or "").strip()
        if code:
            return code
        sid = getattr(r, "shift_id", None)
        if sid is None:
            return ""
        sd = by_id.get(int(sid))
        return sd.shift_code if sd else ""

    for r in rows:
        eid = int(r.employee_id)
        dstr = str(r.work_date)
        matrix.setdefault(eid, {})[dstr] = _row_value(r)

    # Select employee
    emp_raw = (request.values.get("emp") or request.args.get("emp") or "").strip()
    selected_emp_id: Optional[int] = None
    if emp_raw.isdigit():
        selected_emp_id = int(emp_raw)
    elif employees:
        selected_emp_id = int(employees[0].employee_id)

    # Options for selects
    def _hhmm(t):
        return t.strftime("%H:%M") if t else ""

    def _rng(sd):
        return f"{_hhmm(sd.start_time)}–{_hhmm(sd.end_time)}"

    options: List[Tuple[str, str]] = [("", "—")]
    options += [
        ("DES", "Descanso"),
        ("INC", "Incapacidad"),
        ("VAC", "Vacaciones"),
        ("HB", "Habilidad"),
        ("1P", "1/2 Permiso (sin turno)"),
        ("2P", "Permiso (día completo)"),
    ]

    for sd in shift_defs:
        code = (sd.shift_code or "").strip()
        if not code:
            continue
        label = _rng(sd)
        options.append((code, label))
        options.append((f"1P/{code}", f"1/2 · {label}"))

    label_map = {v: l for v, l in options}

    # Calendar event mapping for selected employee
    def _minutes(t):
        if not t:
            return None
        return int(t.hour) * 60 + int(t.minute)

    def _event_for_value(v: str):
        v = (v or "").strip()
        if not v:
            return None
        vup = v.upper()

        # Specials: full-day blocks
        if vup in ("DES", "VAC", "INC", "HB", "2P"):
            cls = "des" if vup == "DES" else ("vac" if vup == "VAC" else ("inc" if vup == "INC" else "spec"))
            return {
                "label": label_map.get(v, v),
                "start": 0,
                "end": 24 * 60,
                "cls": cls,
            }

        if vup == "1P":
            return {"label": label_map.get(v, v), "start": 0, "end": 12 * 60, "cls": "spec"}

        half = False
        code = v
        if vup.startswith("1P/"):
            half = True
            code = v.split("/", 1)[1].strip()

        sd = by_code.get(code)
        if not sd:
            return {"label": v, "start": 0, "end": 0, "cls": "spec"}

        smin = _minutes(sd.start_time) or 0
        emin = _minutes(sd.end_time) or 0
        if emin <= smin:
            emin += 24 * 60

        if half:
            dur = emin - smin
            emin = smin + max(30, dur // 2)

        return {
            "label": label_map.get(v, _rng(sd)),
            "start": smin,
            "end": emin,
            "cls": "half" if half else "shift",
        }

    events_by_date: Dict[str, Dict] = {}
    if selected_emp_id is not None:
        per = matrix.get(selected_emp_id, {})
        for d in days:
            dstr = str(d)
            ev = _event_for_value(per.get(dstr, ""))
            if ev:
                events_by_date[dstr] = ev

    # Save
    if request.method == "POST":
        if selected_emp_id is None:
            flash("Selecciona un empleado.", "warning")
            return redirect(url_for("modulos.turnos_planificacion", group=group, start=str(ws)))

        for d in days:
            dstr = str(d)
            key = f"cell_{selected_emp_id}_{dstr}"
            val = (request.form.get(key) or "").strip()
            try:
                roster.upsert_roster_day(week_id, selected_emp_id, d, val, getattr(current_user, "user_id", None))
            except Exception as ex:
                flash(f"No se pudo guardar {dstr}: {ex}", "danger")
                return redirect(url_for("modulos.turnos_planificacion", group=group, start=str(ws), emp=selected_emp_id))

        flash("Planificación guardada.", "success")
        return redirect(url_for("modulos.turnos_planificacion", group=group, start=str(ws), emp=selected_emp_id))

    return render_template(
        "modulos/turnos_planificacion.html",
        group=group,
        week_start=ws,
        prev_start=str(ws - timedelta(days=7)),
        next_start=str(ws + timedelta(days=7)),
        days=days,
        employees=employees,
        selected_emp_id=selected_emp_id,
        matrix=matrix,
        options=options,
        label_map=label_map,
        events_by_date=events_by_date,
        has_shift_defs=bool(shift_defs),
    )

# -------------------------
# Asistencia (import + vista)
# -------------------------
def _parse_attendance_excel(file_path: str) -> List[Dict]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # detectar cabecera (primera fila)
    header = []
    for cell in ws[1]:
        header.append(str(cell.value).strip().lower() if cell.value is not None else "")
    header_map = {name: idx for idx, name in enumerate(header)}

    def find_col(patterns):
        for name, idx in header_map.items():
            for p in patterns:
                if p in name:
                    return idx
        return None

    col_doc = find_col(["cedula", "cédula", "documento", "doc", "identific", "cc", "c.c", "dni", "id"])
    col_date = find_col(["fecha", "date"])
    col_in = find_col(["entrada", "first", "ingreso", "in"])
    col_out = find_col(["salida", "last", "egreso", "out"])
    col_min = find_col(["min", "minutos", "total", "tiempo", "duracion", "duración"])

    if col_doc is None or col_date is None:
        raise ValueError("No se encontró columna de Cédula/Documento y/o Fecha en el Excel (revisa encabezados).")

    grouped: Dict[Tuple[str, date], Dict] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        doc = _parse_doc_number(r[col_doc] if col_doc < len(r) else None)
        wdate = _parse_date(r[col_date] if col_date < len(r) else None)
        if not doc or not wdate:
            continue

        fin = _parse_time(r[col_in] if (col_in is not None and col_in < len(r)) else None)
        fout = _parse_time(r[col_out] if (col_out is not None and col_out < len(r)) else None)
        mins = None
        if col_min is not None and col_min < len(r):
            try:
                mins = int(float(r[col_min])) if r[col_min] is not None else None
            except Exception:
                mins = None

        key = (doc, wdate)
        entry = grouped.get(key)
        raw = {"row": [("" if x is None else f"{x}") for x in r]}

        if entry is None:
            grouped[key] = {
                "doc_number": doc,
                "work_date": wdate,
                "first_in": fin,
                "last_out": fout,
                "total_minutes": mins,
                "raw": [raw],
            }
        else:
            if fin and (entry["first_in"] is None or fin < entry["first_in"]):
                entry["first_in"] = fin
            if fout and (entry["last_out"] is None or fout > entry["last_out"]):
                entry["last_out"] = fout
            # si viene minutes, dejamos el máximo
            if mins is not None:
                if entry["total_minutes"] is None or mins > entry["total_minutes"]:
                    entry["total_minutes"] = mins
            entry["raw"].append(raw)

    rows = []
    for v in grouped.values():
        if v["total_minutes"] is None and v["first_in"] and v["last_out"]:
            try:
                v["total_minutes"] = max(0, _diff_minutes(v["first_in"], v["last_out"]))
            except Exception:
                v["total_minutes"] = None
        v["raw_text"] = json.dumps(v["raw"], ensure_ascii=False)
        rows.append(v)

    return rows

@modulos_bp.route("/asistencia")
@login_required
def asistencia():
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    y = int(request.args.get("year", date.today().year))
    m = int(request.args.get("month", date.today().month))
    d_from, d_to = _month_range(y, m)

    # fechas del mes
    days = [d_from + timedelta(days=i) for i in range((d_to - d_from).days + 1)]

    empleados = fetch_all(
        "SELECT employee_id, doc_number, first_name, last_name FROM rrhh.hr_employee WHERE is_active=1 ORDER BY last_name, first_name"
    )

    att_rows = fetch_all(
        "SELECT employee_id, work_date, first_in, last_out, total_minutes, has_manual_override "
        "FROM rrhh.vw_attendance_effective "
        "WHERE work_date BETWEEN ? AND ?",
        (d_from, d_to),
    )
    att_map = {(r.employee_id, r.work_date): r for r in att_rows}

    # WFH (si existe)
    wfh_exists = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='wfh_day'"
    ) is not None
    wfh_map = set()
    if wfh_exists:
        wfh_rows = fetch_all(
            "SELECT employee_id, work_date FROM rrhh.wfh_day WHERE is_active=1 AND work_date BETWEEN ? AND ?",
            (d_from, d_to),
        )
        wfh_map = {(r.employee_id, r.work_date) for r in wfh_rows}

    # time reduction (si existe y está aprobado)
    tr_exists = fetch_one(
        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='time_reduction_request'"
    ) is not None
    reduction_map = {}
    if tr_exists:
        tr_rows = fetch_all(
            "SELECT r.employee_id, t.reduction_date, t.minutes "
            "FROM rrhh.wf_request r "
            "JOIN rrhh.time_reduction_request t ON t.request_id = r.request_id "
            "WHERE r.request_type = 'TIME_REDUCTION' AND r.status = 'APPROVED' "
            "AND t.reduction_date BETWEEN ? AND ?",
            (d_from, d_to),
        )
        for rr in tr_rows:
            reduction_map[(rr.employee_id, rr.reduction_date)] = int(rr.minutes or 60)

    # shift assignments (traer los que cruzan el mes)
    shifts = fetch_all(
        "SELECT sa.employee_id, sa.valid_from, sa.valid_to, sd.shift_code, sd.start_time, sd.end_time "
        "FROM rrhh.shift_assignment sa "
        "JOIN rrhh.shift_definition sd ON sd.shift_id = sa.shift_id "
        "WHERE sa.valid_from <= ? AND (sa.valid_to IS NULL OR sa.valid_to >= ?) "
        "ORDER BY sa.employee_id, sa.valid_from",
        (d_to, d_from),
    )
    shifts_by_emp: Dict[int, List] = {}
    for s in shifts:
        shifts_by_emp.setdefault(s.employee_id, []).append(s)

    def shift_for(emp_id: int, day: date):
        lst = shifts_by_emp.get(emp_id, [])
        best = None
        for s in lst:
            if s.valid_from <= day and (s.valid_to is None or s.valid_to >= day):
                best = s
        return best

    matrix = []
    for e in empleados:
        row_days = []
        summary = {"ok": 0, "faltante": 0, "incompleto": 0, "casa": 0, "sin_turno": 0}
        for d in days:
            if d.weekday() >= 5:  # sábado/domingo
                row_days.append({"status": "—", "title": "Fin de semana", "link": None})
                continue

            sh = shift_for(e.employee_id, d)
            if not sh:
                summary["sin_turno"] += 1
                row_days.append({"status": "ST", "title": "Sin turno asignado", "link": None})
                continue

            required = _diff_minutes(sh.start_time, sh.end_time)
            required -= reduction_map.get((e.employee_id, d), 0)
            required = max(0, required)

            att = att_map.get((e.employee_id, d))
            if att is None or att.total_minutes is None:
                if (e.employee_id, d) in wfh_map:
                    summary["casa"] += 1
                    row_days.append({"status": "CASA", "title": f"Trabajo en casa. Requerido {required} min", "link": url_for("modulos.asistencia_ajuste", employee_id=e.employee_id, work_date=d.isoformat())})
                else:
                    summary["faltante"] += 1
                    row_days.append({"status": "F", "title": f"Faltante. Requerido {required} min", "link": url_for("modulos.asistencia_ajuste", employee_id=e.employee_id, work_date=d.isoformat())})
                continue

            mins = int(att.total_minutes or 0)
            ok = mins >= required
            if ok:
                summary["ok"] += 1
                st = "OK*" if att.has_manual_override else "OK"
                row_days.append({"status": st, "title": f"{mins} min (req {required}). In {_time_to_str(att.first_in)} / Out {_time_to_str(att.last_out)}", "link": url_for("modulos.asistencia_ajuste", employee_id=e.employee_id, work_date=d.isoformat())})
            else:
                summary["incompleto"] += 1
                row_days.append({"status": "I", "title": f"Incompleto: {mins} min (req {required}). In {_time_to_str(att.first_in)} / Out {_time_to_str(att.last_out)}", "link": url_for("modulos.asistencia_ajuste", employee_id=e.employee_id, work_date=d.isoformat())})

        # turno actual para mostrar
        sh_today = shift_for(e.employee_id, date.today())
        shift_label = sh_today.shift_code if sh_today else "—"
        matrix.append({
            "employee_id": e.employee_id,
            "doc_number": e.doc_number,
            "name": f"{e.first_name} {e.last_name}",
            "shift": shift_label,
            "days": row_days,
            "summary": summary,
        })

    return render_template("modulos/asistencia.html", year=y, month=m, days=days, matrix=matrix)

@modulos_bp.route("/asistencia/cargar", methods=["GET", "POST"])
@login_required
def asistencia_cargar():
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    if request.method == "POST":
        year_no = int(request.form.get("year_no"))
        month_no = int(request.form.get("month_no"))
        f = request.files.get("file")

        if not f or not f.filename:
            flash("Debes seleccionar un archivo.", "warning")
            return render_template("modulos/asistencia_cargar.html", year_no=year_no, month_no=month_no)

        try:
            meta = save_upload(f, prefix="asistencia")
            checksum = _checksum_file(meta["storage_path"])

            # registrar archivo
            execute(
                "INSERT INTO rrhh.sys_attachment(file_name, mime_type, size_bytes, storage_path, uploaded_by_user_id) "
                "VALUES (?, ?, ?, ?, ?)",
                (meta["file_name"], f.mimetype or "application/octet-stream", meta["size_bytes"], meta["storage_path"], current_user.user_id),
            )
            file_id = execute_scalar("SELECT SCOPE_IDENTITY()")

            # batch
            execute(
                "INSERT INTO rrhh.att_import_batch(year_no, month_no, file_name, file_id, checksum, uploaded_by_user_id) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (year_no, month_no, meta["file_name"], file_id, checksum, current_user.user_id),
            )
            batch_id = int(execute_scalar("SELECT SCOPE_IDENTITY()"))

            # parse excel (todos los registros)
            parsed = _parse_attendance_excel(meta["storage_path"])
            doc_col = _att_import_doc_column()

            # insert rows (OK / ERROR si no existe empleado)
            for r in parsed:
                # validar empleado existe
                emp = fetch_one("SELECT employee_id FROM rrhh.hr_employee WHERE doc_number = ? AND is_active=1", (r["doc_number"],))
                if not emp:
                    execute(
                        f"INSERT INTO rrhh.att_import_row(batch_id, {doc_col}, work_date, first_in, last_out, total_minutes, raw_text, load_status, error_message) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, 'ERROR', ?)",
                        (batch_id, r["doc_number"], r["work_date"], r["first_in"], r["last_out"], r["total_minutes"], r["raw_text"], "Empleado no existe (doc_number no encontrado)"),
                    )
                else:
                    execute(
                        f"INSERT INTO rrhh.att_import_row(batch_id, {doc_col}, work_date, first_in, last_out, total_minutes, raw_text, load_status, error_message) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, 'OK', NULL)",
                        (batch_id, r["doc_number"], r["work_date"], r["first_in"], r["last_out"], r["total_minutes"], r["raw_text"]),
                    )

            # aplicar a att_day
            call_proc("rrhh.sp_apply_attendance_batch", [batch_id, current_user.user_id])

            flash(f"Asistencia cargada y aplicada. Batch #{batch_id}.", "success")
            return redirect(url_for("modulos.asistencia", year=year_no, month=month_no))

        except Exception as ex:
            flash(f"Error cargando asistencia: {ex}", "danger")

    today = date.today()
    return render_template("modulos/asistencia_cargar.html", year_no=today.year, month_no=today.month)

@modulos_bp.route("/asistencia/ajuste", methods=["GET", "POST"])
@login_required
def asistencia_ajuste():
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    employee_id = int(request.values.get("employee_id"))
    work_date_s = request.values.get("work_date")
    work_date = datetime.strptime(work_date_s, "%Y-%m-%d").date()

    emp = fetch_one("SELECT employee_id, doc_number, first_name, last_name FROM rrhh.hr_employee WHERE employee_id=?", (employee_id,))
    if not emp:
        flash("Empleado no encontrado.", "warning")
        return redirect(url_for("modulos.asistencia"))

    # asistencia efectiva
    att = fetch_one(
        "SELECT employee_id, work_date, first_in, last_out, total_minutes, has_manual_override "
        "FROM rrhh.vw_attendance_effective WHERE employee_id=? AND work_date=?",
        (employee_id, work_date),
    )

    # turno del día y requerido
    sh = fetch_one(
        "SELECT TOP (1) sd.shift_code, sd.start_time, sd.end_time "
        "FROM rrhh.shift_assignment sa "
        "JOIN rrhh.shift_definition sd ON sd.shift_id = sa.shift_id "
        "WHERE sa.employee_id=? AND sa.valid_from <= ? AND (sa.valid_to IS NULL OR sa.valid_to >= ?) "
        "ORDER BY sa.valid_from DESC",
        (employee_id, work_date, work_date),
    )
    required = _diff_minutes(sh.start_time, sh.end_time) if sh else None

    if request.method == "POST":
        reason = (request.form.get("reason") or "").strip()
        comment = (request.form.get("comment") or "").strip() or None
        first_in = _parse_time(request.form.get("first_in") or None)
        last_out = _parse_time(request.form.get("last_out") or None)
        total_minutes = request.form.get("total_minutes")
        total_minutes = int(total_minutes) if total_minutes not in (None, "",) else None

        # quick action: cumplir
        if request.form.get("action") == "cumplio":
            if required is None:
                flash("No hay turno asignado para calcular requerido.", "warning")
                return redirect(url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=work_date.isoformat()))
            total_minutes = required
            if not reason:
                reason = "NOVEDAD: cumplimiento (salida temprana u otra)"

        if not reason:
            flash("Debes indicar un motivo (reason).", "warning")
            return redirect(url_for("modulos.asistencia_ajuste", employee_id=employee_id, work_date=work_date.isoformat()))

        try:
            # desactivar overrides anteriores
            execute(
                "UPDATE rrhh.att_manual_override SET is_active=0 "
                "WHERE employee_id=? AND work_date=? AND is_active=1",
                (employee_id, work_date),
            )
            execute(
                "INSERT INTO rrhh.att_manual_override(employee_id, work_date, first_in, last_out, total_minutes, reason, comment, created_by_user_id, is_active) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)",
                (employee_id, work_date, first_in, last_out, total_minutes, reason, comment, current_user.user_id),
            )
            flash("Ajuste/Novedad guardado.", "success")
            return redirect(url_for("modulos.asistencia", year=work_date.year, month=work_date.month))
        except Exception as ex:
            flash(f"No se pudo guardar el ajuste: {ex}", "danger")

    return render_template(
        "modulos/asistencia_ajuste.html",
        emp=emp,
        work_date=work_date,
        att=att,
        shift=sh,
        required_minutes=required,
    )

# -------------------------
# Trabajo en casa (aprobaciones)
# -------------------------

def _is_admin_or_rrhh() -> bool:
    roles = getattr(current_user, 'roles', None) or []
    return bool(
        getattr(current_user, 'is_admin', False)
        or (ROLE_ADMIN in roles)
        or (ROLE_RRHH in roles)
        or ('ADMINISTRADOR' in roles)
        or ('RRHH' in roles)
    )


def _wfh_week_range(d: date) -> tuple[date, date]:
    week_start = d - timedelta(days=d.weekday())
    return week_start, week_start + timedelta(days=6)


def _wfh_week_has_weekday_holiday(work_date: date) -> bool:
    """True si existe un festivo entre semana (lun-vie) en la semana de work_date."""
    try:
        has_table = (
            fetch_one(
                "SELECT 1 FROM INFORMATION_SCHEMA.TABLES "
                "WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='hr_holiday'"
            )
            is not None
        )
    except Exception:
        return False

    if not has_table:
        return False

    week_start, week_end = _wfh_week_range(work_date)

    holiday_has_is_active = (
        fetch_one(
            "SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS "
            "WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='hr_holiday' AND COLUMN_NAME='is_active'"
        )
        is not None
    )

    def _query(active_only: bool):
        sql = "SELECT holiday_date FROM rrhh.hr_holiday WHERE holiday_date BETWEEN ? AND ?"
        params = [week_start, week_end]
        if active_only and holiday_has_is_active:
            sql += " AND is_active=1"
        return fetch_all(sql, tuple(params)) or []

    hols = _query(True)
    if not hols and holiday_has_is_active:
        hols = _query(False)

    for h in hols:
        try:
            if h.holiday_date and h.holiday_date.weekday() < 5:
                return True
        except Exception:
            continue
    return False


def _validate_wfh_rules(employee_id: int, work_date: date) -> tuple[bool, str]:
    """
    Reglas:
      1) No se permiten 2 días seguidos.
      2) Semanas con festivo (lun-vie) solo permiten 1 día de trabajo en casa.
    """

    prev_day = work_date - timedelta(days=1)
    next_day = work_date + timedelta(days=1)
    adj = fetch_one(
        "SELECT TOP (1) work_date AS d "
        "FROM rrhh.wfh_day "
        "WHERE employee_id=? AND work_date IN (?, ?) AND work_date <> ?",
        (int(employee_id), prev_day, next_day, work_date),
    )
    if adj:
        return (False, "No puedes solicitar Trabajo en casa en días consecutivos. Revisa el día anterior/siguiente.")

    if _wfh_week_has_weekday_holiday(work_date):
        week_start, week_end = _wfh_week_range(work_date)
        c_row = fetch_one(
            "SELECT COUNT(1) AS c "
            "FROM rrhh.wfh_day "
            "WHERE employee_id=? AND work_date BETWEEN ? AND ? AND work_date <> ?",
            (int(employee_id), week_start, week_end, work_date),
        )
        if c_row and int(getattr(c_row, 'c', 0) or 0) >= 1:
            return (False, "En semanas con festivo entre semana solo se permite 1 día de Trabajo en casa.")

    return True, ""


@modulos_bp.route("/trabajo-casa")
@login_required
def trabajo_casa():
    """Vista RRHH/ADMIN: registros aprobados + registrar manual."""
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    exists = (
        fetch_one(
            "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='wfh_day'"
        )
        is not None
    )
    if not exists:
        flash("Tabla rrhh.wfh_day no existe. Aplica el script SQL para crearla.", "warning")
        return render_template("modulos/trabajo_casa.html", registros=[], empleados=[])

    registros = fetch_all(
        "SELECT w.employee_id, w.work_date, w.reason, e.doc_number, e.first_name, e.last_name "
        "FROM rrhh.wfh_day w JOIN rrhh.hr_employee e ON e.employee_id = w.employee_id "
        "WHERE w.is_active=1 ORDER BY w.work_date DESC"
    )
    empleados = fetch_all(
        "SELECT employee_id, doc_number, first_name, last_name "
        "FROM rrhh.hr_employee WHERE is_active=1 ORDER BY last_name, first_name"
    )
    return render_template("modulos/trabajo_casa.html", registros=registros, empleados=empleados)


@modulos_bp.route("/trabajo-casa/nuevo", methods=["POST"])
@login_required
def trabajo_casa_nuevo():
    if not _require_admin():
        return redirect(url_for("modulos.dashboard"))

    employee_id = int(request.form.get("employee_id"))
    work_date = datetime.strptime(request.form.get("work_date"), "%Y-%m-%d").date()
    reason = (request.form.get("reason") or "Trabajo en casa").strip()

    ok, msg = _validate_wfh_rules(int(employee_id), work_date)
    if not ok:
        flash(msg, "warning")
        return redirect(url_for("modulos.trabajo_casa"))

    try:
        execute(
            "MERGE rrhh.wfh_day AS t "
            "USING (SELECT ? AS employee_id, ? AS work_date) AS s "
            "ON (t.employee_id = s.employee_id AND t.work_date = s.work_date) "
            "WHEN MATCHED THEN UPDATE SET is_active=1, reason=? "
            "WHEN NOT MATCHED THEN INSERT (employee_id, work_date, reason, created_by_user_id, is_active) "
            "VALUES (?, ?, ?, ?, 1);",
            (employee_id, work_date, reason, employee_id, work_date, reason, current_user.user_id),
        )
        flash("Día de trabajo en casa registrado.", "success")
    except Exception as ex:
        flash(f"No se pudo registrar: {ex}", "danger")

    return redirect(url_for("modulos.trabajo_casa"))


@modulos_bp.route("/trabajo-casa/aprobaciones")
@login_required
def trabajo_casa_aprobaciones():
    """Bandeja de aprobaciones WFH."""

    is_backoffice = _is_admin_or_rrhh()
    my_emp_id = getattr(current_user, "employee_id", None)

    if not my_emp_id:
        flash("Tu usuario no está asociado a un empleado. Contacta a RRHH.", "error")
        return redirect(url_for("modulos.dashboard"))

    if not is_backoffice:
        if not manager_has_subordinates(int(my_emp_id)):
            flash("No tienes solicitudes de Trabajo en casa para aprobar.", "warning")
            return redirect(url_for("modulos.dashboard"))

    exists = fetch_one(
        "SELECT 1 AS ok FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='rrhh' AND TABLE_NAME='wfh_day'"
    )
    if not exists:
        flash("No existe la tabla rrhh.wfh_day. Revisa el script de base de datos.", "error")
        return redirect(url_for("modulos.dashboard"))

    if is_backoffice:
        pendientes = fetch_all(
            """
            SELECT
              e.employee_id,
              (e.first_name + ' ' + e.last_name) AS employee_name,
              e.doc_number,
              w.work_date,
              w.reason,
              w.is_active,
              w.created_at
            FROM rrhh.wfh_day w
            JOIN rrhh.hr_employee e ON e.employee_id = w.employee_id
            WHERE w.is_active = 0
            ORDER BY w.created_at DESC
            """
        )

        aprobadas = fetch_all(
            """
            SELECT
              e.employee_id,
              (e.first_name + ' ' + e.last_name) AS employee_name,
              e.doc_number,
              w.work_date,
              w.reason,
              w.is_active,
              w.created_at
            FROM rrhh.wfh_day w
            JOIN rrhh.hr_employee e ON e.employee_id = w.employee_id
            WHERE w.is_active = 1
              AND w.created_at >= DATEADD(day, -60, GETDATE())
            ORDER BY w.created_at DESC
            """
        )

    else:
        pendientes = fetch_all(
            """
            SELECT
              e.employee_id,
              (e.first_name + ' ' + e.last_name) AS employee_name,
              e.doc_number,
              w.work_date,
              w.reason,
              w.is_active,
              w.created_at
            FROM rrhh.wfh_day w
            JOIN rrhh.hr_employee e ON e.employee_id = w.employee_id
            JOIN rrhh.hr_employee_manager mm
              ON mm.employee_id = w.employee_id
             AND mm.is_primary = 1
             AND mm.manager_employee_id = ?
             AND mm.valid_from <= w.work_date
             AND (mm.valid_to IS NULL OR mm.valid_to >= w.work_date)
            WHERE w.is_active = 0
              AND w.employee_id <> ?
            ORDER BY w.created_at DESC
            """,
            (int(my_emp_id), int(my_emp_id)),
        )

        aprobadas = fetch_all(
            """
            SELECT
              e.employee_id,
              (e.first_name + ' ' + e.last_name) AS employee_name,
              e.doc_number,
              w.work_date,
              w.reason,
              w.is_active,
              w.created_at
            FROM rrhh.wfh_day w
            JOIN rrhh.hr_employee e ON e.employee_id = w.employee_id
            JOIN rrhh.hr_employee_manager mm
              ON mm.employee_id = w.employee_id
             AND mm.is_primary = 1
             AND mm.manager_employee_id = ?
             AND mm.valid_from <= w.work_date
             AND (mm.valid_to IS NULL OR mm.valid_to >= w.work_date)
            WHERE w.is_active = 1
              AND w.created_at >= DATEADD(day, -60, GETDATE())
              AND w.employee_id <> ?
            ORDER BY w.created_at DESC
            """,
            (int(my_emp_id), int(my_emp_id)),
        )

    return render_template(
        "modulos/trabajo_casa_aprobaciones.html",
        pendientes=pendientes or [],
        aprobadas=aprobadas or [],
        is_backoffice=is_backoffice,
    )


@modulos_bp.route("/trabajo-casa/aprobar", methods=["POST"])
@login_required
def trabajo_casa_aprobar():
    employee_id = int(request.form.get("employee_id"))
    work_date = datetime.strptime(request.form.get("work_date"), "%Y-%m-%d").date()

    # ADMIN/RRHH aprueban todo; jefe aprueba su equipo
    if not _is_admin_or_rrhh():
        mgr_id = get_manager_for_employee(int(employee_id), work_date)
        my_emp_id = getattr(current_user, "employee_id", None)
        if not my_emp_id or not mgr_id or int(mgr_id) != int(my_emp_id):
            flash("No tienes permiso para aprobar esta solicitud.", "warning")
            return redirect(url_for("modulos.trabajo_casa_aprobaciones"))

    ok, msg = _validate_wfh_rules(int(employee_id), work_date)
    if not ok:
        flash(f"No se puede aprobar: {msg}", "warning")
        return redirect(url_for("modulos.trabajo_casa_aprobaciones"))

    try:
        rc = execute(
            "UPDATE rrhh.wfh_day SET is_active=1 WHERE employee_id=? AND work_date=? AND is_active=0",
            (employee_id, work_date),
        )
        if rc:
            flash("Solicitud aprobada.", "success")
        else:
            flash("La solicitud no estaba pendiente o no existe.", "warning")
    except Exception as ex:
        flash(f"No se pudo aprobar: {ex}", "danger")

    return redirect(url_for("modulos.trabajo_casa_aprobaciones"))


@modulos_bp.route("/trabajo-casa/rechazar", methods=["POST"])
@login_required
def trabajo_casa_rechazar():
    employee_id = int(request.form.get("employee_id"))
    work_date = datetime.strptime(request.form.get("work_date"), "%Y-%m-%d").date()

    if not _is_admin_or_rrhh():
        mgr_id = get_manager_for_employee(int(employee_id), work_date)
        my_emp_id = getattr(current_user, "employee_id", None)
        if not my_emp_id or not mgr_id or int(mgr_id) != int(my_emp_id):
            flash("No tienes permiso para rechazar esta solicitud.", "warning")
            return redirect(url_for("modulos.trabajo_casa_aprobaciones"))

    try:
        rc = execute(
            "DELETE FROM rrhh.wfh_day WHERE employee_id=? AND work_date=? AND is_active=0",
            (employee_id, work_date),
        )
        if rc:
            flash("Solicitud rechazada.", "success")
        else:
            flash("La solicitud no estaba pendiente o no existe.", "warning")
    except Exception as ex:
        flash(f"No se pudo rechazar: {ex}", "danger")

    return redirect(url_for("modulos.trabajo_casa_aprobaciones"))


 

def _safe_import(name: str) -> None:
    try:
        __import__(name, fromlist=['*'])
    except AssertionError:
        # endpoint duplicado; ignorar
        pass
    except SyntaxError:
        # módulo con error de sintaxis; ignorar para no tumbar el arranque
        pass
    except Exception:
        pass

_safe_import('blueprints.modulos_chequera')
_safe_import('blueprints.modulos_hora_flexible')
_safe_import('blueprints.modulos_incapacidad')
_safe_import('blueprints.modulos_reportes')
_safe_import('blueprints.modulos_vacaciones')
