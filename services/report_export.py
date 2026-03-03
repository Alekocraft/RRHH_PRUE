from __future__ import annotations

import io
from datetime import datetime
from typing import Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

# Traducción de estados técnicos (workflow) a etiquetas en español para reportes (UI/export).
_STATUS_ES = {
    "DRAFT": "Borrador",
    "SUBMITTED": "Enviada",
    "PENDING": "Pendiente",
    "APPROVED": "Aprobada",
    "REJECTED": "Rechazada",
    "SKIPPED": "Omitida",
    "CANCELLED": "Cancelada",
    "CANCELED": "Cancelada",
    # Variantes ya guardadas en español (normalizamos)
    "APROBADO": "Aprobado",
    "APROBADA": "Aprobada",
    "RECHAZADO": "Rechazado",
    "RECHAZADA": "Rechazada",
    "PENDIENTE": "Pendiente",
    "CANCELADO": "Cancelado",
    "CANCELADA": "Cancelada",
    # Vacaciones (si llegan en reportes)
    "PENDING_MANAGER": "Pendiente jefe",
    "PENDING_HR": "Pendiente RRHH",
    "REJECTED_MANAGER": "Rechazada jefe",
    "REJECTED_HR": "Rechazada RRHH",
    "REJECTED_JEFE": "Rechazada jefe",
}


def _safe_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime,)):
        return v.strftime("%Y-%m-%d %H:%M")
    # Traduce estados del workflow cuando vienen como string
    try:
        s = f"{v}"
    except Exception:
        return ""
    up = s.strip().upper()
    if up in _STATUS_ES:
        return _STATUS_ES[up]
    return s


def build_excel(
    title: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
) -> bytes:
    """Build an XLSX payload.

    Fix: avoid iterating ws.columns when row 1 has merged cells (openpyxl returns MergedCell),
    which lacks column_letter. We compute column widths from headers + streamed rows instead.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    header_list = list(headers) if headers else [""]

    # Title row (merged across header columns)
    ws.append([title])
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=max(1, len(header_list)),
    )
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left")

    # Header row
    ws.append(header_list)
    for cell in ws[2]:
        cell.font = Font(bold=True)

    # Track max lengths per column for width auto-fit
    col_count = max(1, len(header_list))
    max_lens = [0] * col_count

    # Title contributes to first column width
    max_lens[0] = max(max_lens[0], len(f"{title or ""}"))

    for i, h in enumerate(header_list[:col_count]):
        max_lens[i] = max(max_lens[i], len(f"{h or ""}"))

    # Data rows
    for r in rows:
        row_vals = [_safe_str(x) for x in r]
        ws.append(row_vals)

        # In case a row has more columns than headers, extend width tracking + keep sheet usable
        if len(row_vals) > col_count:
            max_lens.extend([0] * (len(row_vals) - col_count))
            col_count = len(row_vals)

        for i in range(min(col_count, len(row_vals))):
            v = row_vals[i]
            max_lens[i] = max(max_lens[i], len(f"{v or ""}"))

    # Apply widths (bounded)
    for i in range(col_count):
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = min(max(10, max_lens[i] + 2), 45)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()



def build_pdf(
    title: str,
    subtitle_lines: List[str],
    headers: Sequence[str],
    rows: Iterable[Sequence],
) -> bytes:
    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=18,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()

    story = [Paragraph(title, styles["Title"]), Spacer(1, 10)]
    for line in subtitle_lines or []:
        story.append(Paragraph(line, styles["Normal"]))
    if subtitle_lines:
        story.append(Spacer(1, 10))

    data = [list(headers)]
    for r in rows:
        data.append([_safe_str(x) for x in r])

    tbl = Table(data, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#A73493")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (-1, 0), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ]
        )
    )

    story.append(tbl)
    doc.build(story)
    return bio.getvalue()
