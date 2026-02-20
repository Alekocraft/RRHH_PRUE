import os
import csv
from typing import Dict, Tuple, List, Optional
from openpyxl import load_workbook

def _norm(s: str) -> str:
    return (s or "").strip().lower()

def _guess_col(cols: List[str], candidates: List[str]) -> Optional[str]:
    cols_l = {_norm(c): c for c in cols}
    for cand in candidates:
        if _norm(cand) in cols_l:
            return cols_l[_norm(cand)]
    # fuzzy: contains
    for c in cols:
        cl=_norm(c)
        for cand in candidates:
            if _norm(cand) in cl:
                return c
    return None

def _read_xlsx(path: str) -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data: List[Dict[str, str]] = []
    for r in rows[1:]:
        rec: Dict[str, str] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = r[i] if i < len(r) else None
            rec[h] = "" if val is None else str(val).strip()
        if any(rec.values()):
            data.append(rec)
    return data

def _read_csv(path: str) -> List[Dict[str, str]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [{k: (v or "").strip() for k,v in row.items()} for row in reader]

def load_org_mapping() -> Tuple[Dict[str, Dict], List[str], List[str]]:
    """Lee el excel/csv de estructura (jefe/área).
    Configurar env: RRHH_ORG_XLSX_PATH
    Retorna: (mapping_por_usuario, lista_jefes, lista_areas)
    """
    path = os.getenv("RRHH_ORG_XLSX_PATH", "").strip()
    if not path or not os.path.exists(path):
        return {}, [], []

    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        data = _read_xlsx(path)
    elif ext == ".csv":
        data = _read_csv(path)
    else:
        return {}, [], []

    if not data:
        return {}, [], []

    cols = list(data[0].keys())

    col_user = _guess_col(cols, ["usuario","user","username","samaccountname","login","cuenta","ad_username","dominio","correo","email"])
    col_mgr  = _guess_col(cols, ["jefe","manager","supervisor","lider","líder","responsable"])
    col_area = _guess_col(cols, ["area","área","departamento","department","dependencia","unidad"])

    if not col_user:
        return {}, [], []

    mapping: Dict[str, Dict] = {}
    bosses=set()
    areas=set()

    for row in data:
        u = (row.get(col_user) or "").strip()
        if not u:
            continue

        u_key = u.split("\\")[-1].split("@")[0].lower()

        mgr = (row.get(col_mgr) or "").strip() if col_mgr else ""
        mgr_key = mgr.split("\\")[-1].split("@")[0].lower() if mgr else ""

        area = (row.get(col_area) or "").strip() if col_area else ""

        mapping[u_key] = {"manager": mgr_key, "area": area}

        if mgr_key:
            bosses.add(mgr_key)
        if area:
            areas.add(area)

    return mapping, sorted(bosses), sorted(areas)
